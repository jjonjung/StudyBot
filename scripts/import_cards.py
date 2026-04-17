"""
import_cards.py  —  StudyBot 카드 분류기 & DB 임포터
======================================================
역할:
  1. card/ 폴더의 xlsx/csv 원본 데이터를 읽는다
  2. Unreal / C++ / CS / Company / Algorithm 카테고리로 분류한다
  3. flashcards_classified.xlsx 를 시트 5개로 저장한다
  4. --db 옵션 추가 시 MySQL studybot.flashcards 에 INSERT

원본 파일:
  card/unreal_interview_qa.xlsx           — Unreal/C++/CS 혼합 (자동분류)
  card/cpp_tech_interview_flashcards.xlsx — C++ 중심 (자동분류)
  card/CS_260325.csv                      — CS 중심 (cp949, 자동분류)
  card/company_interview_qa.xlsx          — 기업 기출 (회사명 컬럼 포함)
  card/algorithm_interview_qa.xlsx        — 알고리즘 (코드·복잡도 포함)

사용:
  pip install openpyxl mysql-connector-python
  python scripts/import_cards.py          # xlsx만 생성
  python scripts/import_cards.py --db     # xlsx + DB INSERT
"""

import sys, os, csv, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE     = os.path.join(os.path.dirname(__file__), "..")
CARD_DIR = os.path.join(BASE, "card")

DB_CONFIG = {
    "host":     os.getenv("DB_HOST",     "localhost"),
    "port":     int(os.getenv("DB_PORT", "3306")),
    "user":     os.getenv("DB_USER",     "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME",     "studybot"),
    "charset":  "utf8mb4",
}

# ── 분류 키워드 ──────────────────────────────────────────
UNREAL_KW = {
    "언리얼","unreal","ue5","ue4","actor","uactor","component","blueprint",
    "uobject","gameinstance","ugameinstance","subsystem","delegate",
    "eventdispatcher","event dispatcher","umg","widget","playercontroller",
    "acharacter","apawn","ufunction","uproperty","linetrace","raycast",
    "primitivecomponent","statetree","tick","deltatime","replication",
    "rpc","multicast","fjsonobject","fjsonserializer","dataasset","hud",
    "uinterface","ability","gameplayability","gamedataasset","channelcast",
}
CPP_KW = {
    "c++","cpp","virtual","vptr","vtable","raii","move semantics","unique_ptr",
    "shared_ptr","weak_ptr","template","stl","vector","unordered_map","lambda",
    "atomic","mutex","thread","memory_order","false sharing","cache locality",
    "bitset","rvalue","r-value","lvalue","복사 생성자","대입 연산자",
    "얕은 복사","깊은 복사","happens-before","erase-remove","alignas",
    "thread pool","pathfinding","race condition","소멸자","가상 함수",
}
CS_KW = {
    "tcp","udp","네트워크","network","os","운영체제","프로세스","스레드",
    "데드락","deadlock","가상 메모리","페이지 폴트","캐시","뮤텍스","세마포어",
    "배열","연결 리스트","해시","스택","큐","힙","bst","트리","그래프",
    "다익스트라","dijkstra","a*","astar","우선순위 큐","시간 복잡도","db","sql",
    "lock","자료구조","알고리즘","정렬","binary",
}

def classify(q: str, a: str, hint: str = "") -> str:
    text = (q + " " + a).lower()
    text = re.sub(r"\s+", " ", text)

    u = sum(1 for k in UNREAL_KW if k in text)
    c = sum(1 for k in CPP_KW    if k in text)
    s = sum(1 for k in CS_KW     if k in text)

    hint_l = hint.lower()
    if hint_l in ("unreal", "언리얼"):             u += 6
    elif hint_l in ("c++", "cpp"):                 c += 6
    elif hint_l in ("cs", "자료구조", "알고리즘"): s += 6

    best = max({"Unreal": u, "C++": c, "CS": s}, key=lambda k: {"Unreal":u,"C++":c,"CS":s}[k])
    return best

# ── 원본 읽기 ────────────────────────────────────────────
def load_unreal_xlsx():
    wb = openpyxl.load_workbook(os.path.join(CARD_DIR, "unreal_interview_qa.xlsx"))
    ws = wb["Unreal_QA"]
    out = []
    for _, q, a in ws.iter_rows(min_row=2, values_only=True):
        if q and a:
            cat = classify(str(q), str(a), hint="Unreal")
            out.append(dict(category=cat, question=str(q).strip(),
                            answer=str(a).strip(), difficulty="Normal"))
    return out

def load_cpp_xlsx():
    wb = openpyxl.load_workbook(os.path.join(CARD_DIR, "cpp_tech_interview_flashcards.xlsx"))
    ws = wb["Sheet1"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q, a = row[0], row[1]
        if q and a:
            cat = classify(str(q), str(a), hint="C++")
            out.append(dict(category=cat, question=str(q).strip(),
                            answer=str(a).strip(), difficulty="Normal"))
    return out

def load_cs_csv():
    out = []
    path = os.path.join(CARD_DIR, "CS_260325.csv")
    with open(path, encoding="cp949", errors="replace") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)
        for row in reader:
            if len(row) < 3 or not row[1].strip():
                continue
            hint, q, a = row[0].strip(), row[1].strip(), row[2].strip()
            cat = classify(q, a, hint=hint)
            out.append(dict(category=cat, question=q, answer=a, difficulty="Normal"))
    return out

# ── xlsx 저장 ─────────────────────────────────────────────
SHEET_STYLE = {
    "Unreal":    dict(header_fill="1F4E79", row_fills=["D6E4F0","EBF5FB"]),
    "C++":       dict(header_fill="1E5631", row_fills=["D5F5E3","EAFAF1"]),
    "CS":        dict(header_fill="7B241C", row_fills=["FADBD8","FDEDEC"]),
    "Company":   dict(header_fill="4A235A", row_fills=["E8DAEF","F5EEF8"]),
    "Algorithm": dict(header_fill="1F3864", row_fills=["D9E1F2","EEF2FB"]),
}
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build_sheet(wb, name: str, cards: list):
    ws = wb.create_sheet(title=name)
    st = SHEET_STYLE[name]
    hfill = PatternFill("solid", fgColor=st["header_fill"])
    hfont = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")

    is_company   = (name == "Company")
    is_algorithm = (name == "Algorithm")

    if is_algorithm:
        headers = ["#", "질문", "핵심 조건", "선택 이유", "모범 답변",
                   "C++ 코드", "C# 코드", "시간복잡도", "난이도"]
        col_w   = [4,   38,   24,      24,       40,
                   48,       48,        14,        10]
    elif is_company:
        headers = ["#", "회사명", "질문", "모범 답변", "난이도"]
        col_w   = [5, 12, 44, 72, 10]
    else:
        headers = ["#", "카테고리", "질문", "모범 답변", "난이도"]
        col_w   = [5, 10, 46, 72, 10]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ri, card in enumerate(cards, 1):
        fill     = PatternFill("solid", fgColor=st["row_fills"][ri % 2])
        row_font = Font(name="맑은 고딕", size=10)
        code_font= Font(name="Consolas",  size=9)

        if is_algorithm:
            vals = [ri,
                    card["question"], card.get("core_conditions",""),
                    card.get("selection_reason",""), card["answer"],
                    card.get("code_cpp",""), card.get("code_csharp",""),
                    card.get("time_complexity",""), card["difficulty"]]
        elif is_company:
            vals = [ri, card.get("company",""), card["question"],
                    card["answer"], card["difficulty"]]
        else:
            vals = [ri, card["category"], card["question"],
                    card["answer"], card["difficulty"]]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri+1, column=ci, value=v)
            c.fill   = fill; c.border = BORDER
            # 코드 컬럼(Algorithm 6,7번)은 Consolas
            c.font   = code_font if (is_algorithm and ci in (6,7)) else row_font
            c.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if ci in (1, len(headers)) else "left",
            )

    for ci, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

def save_xlsx(all_cards: list) -> dict:
    by_cat: dict = {"Unreal": [], "C++": [], "CS": [], "Company": [], "Algorithm": []}
    for card in all_cards:
        cat = card["category"]
        if cat in by_cat:
            by_cat[cat].append(card)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for cat in ("Unreal", "C++", "CS", "Company", "Algorithm"):
        build_sheet(wb, cat, by_cat[cat])

    out = os.path.join(CARD_DIR, "flashcards_classified.xlsx")
    wb.save(out)
    print(f"\n[xlsx] {out}")
    for cat, cards in by_cat.items():
        print(f"  {cat:10s}: {len(cards):2d}장")
    return by_cat


# ── 알고리즘 카드 읽기 ────────────────────────────────────
def load_algorithm_xlsx():
    """card/algorithm_interview_qa.xlsx 읽기
    컬럼: # | 질문 | 핵심조건 | 선택이유 | 모범답변 | C++코드 | C#코드 | 시간복잡도 | 난이도
    """
    path = os.path.join(CARD_DIR, "algorithm_interview_qa.xlsx")
    if not os.path.exists(path):
        print("[경고] algorithm_interview_qa.xlsx 없음 → create_algorithm_cards.py 먼저 실행")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2 or not row[1]: continue
        q    = str(row[1]).strip()
        core = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        sel  = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        ans  = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        cpp  = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        csh  = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        tc   = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        diff = str(row[8]).strip() if len(row) > 8 and row[8] else "Normal"
        if diff not in ("Easy","Normal","Hard"): diff = "Normal"
        out.append(dict(
            category="Algorithm",
            question=q, answer=ans,
            core_conditions=core, selection_reason=sel,
            code_cpp=cpp, code_csharp=csh,
            time_complexity=tc, difficulty=diff,
        ))
    print(f"[algorithm] {len(out)}장 로드")
    return out

# ── 기업 기출 읽기 ────────────────────────────────────────
def load_company_xlsx():
    """card/company_interview_qa.xlsx 읽기
    컬럼: # | 회사명 | 질문 | 모범 답변 | 난이도
    파일이 없으면 빈 리스트 반환 (선택적 데이터)
    """
    path = os.path.join(CARD_DIR, "company_interview_qa.xlsx")
    if not os.path.exists(path):
        print("[경고] company_interview_qa.xlsx 없음, Company 카드 건너뜀")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4: continue
        company = row[1]
        q       = row[2]
        a       = row[3]
        diff    = row[4] if len(row) > 4 else "Normal"
        if not q or not a: continue
        diff = diff if diff in ("Easy", "Normal", "Hard") else "Normal"
        out.append(dict(
            category   = "Company",
            company    = str(company).strip() if company else "",
            question   = str(q).strip(),
            answer     = str(a).strip(),
            difficulty = diff,
        ))
    print(f"[company] {len(out)}장 로드")
    return out

# ── DB INSERT ─────────────────────────────────────────────
def insert_db(by_cat: dict):
    try:
        import mysql.connector
    except ImportError:
        print("[DB] pip install mysql-connector-python 필요"); return
    conn = mysql.connector.connect(**DB_CONFIG)
    cur = conn.cursor()
    total = 0
    for cat, cards in by_cat.items():
        for card in cards:
            cur.execute(
                """CALL sp_cards_create(
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s
                   )""",
                (card["category"],
                 card.get("company")          or None,
                 card["question"],
                 card.get("answer",""),
                 card.get("difficulty","Normal"),
                 card.get("core_conditions")  or None,
                 card.get("selection_reason") or None,
                 card.get("code_cpp")         or None,
                 card.get("code_csharp")      or None,
                 card.get("time_complexity")  or None,
                )
            )
            for result in cur.stored_results():
                result.fetchall()
            total += 1
    conn.commit(); cur.close(); conn.close()
    print(f"[DB] {total}건 INSERT 완료")

if __name__ == "__main__":
    print("=== StudyBot 카드 분류 ===")
    cards  = (load_unreal_xlsx() + load_cpp_xlsx()
              + load_cs_csv() + load_company_xlsx()
              + load_algorithm_xlsx())
    print(f"원본 합계: {len(cards)}장")
    by_cat = save_xlsx(cards)
    if "--db" in sys.argv:
        insert_db(by_cat)
    else:
        print("\n[안내] DB INSERT는 --db 옵션 추가 후 실행하세요.")
