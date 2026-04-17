"""
create_algorithm_cards.py  —  알고리즘 면접 카드 생성기
=========================================================
card/algorithm_interview_qa.xlsx 를 샘플 데이터와 함께 생성합니다.

컬럼:
  # | 질문 | 핵심 조건 | 선택 이유 | 모범 답변 | C++ 코드 | C# 코드 | 시간복잡도 | 난이도

사용:
  python scripts/create_algorithm_cards.py
  → 이후 import_cards.py --db 로 DB에 삽입
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = os.path.join(os.path.dirname(__file__), "..")
OUT_PATH = os.path.join(BASE, "card", "algorithm_interview_qa.xlsx")

THIN   = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 샘플 카드 ──────────────────────────────────────────────
# (질문, 핵심조건, 선택이유, 모범답변, C++코드, C#코드, 시간복잡도, 난이도)
ALGO_CARDS = [
    (
        "정렬된 배열에서 타깃 값의 인덱스를 찾으세요.",
        "배열이 정렬돼 있을 것\n중복 없음(또는 첫 인덱스 반환)\n인덱스 범위 내 탐색",
        "정렬 보장 → 매 반복마다 탐색 범위를 절반으로 줄일 수 있음\n선형 탐색 O(N) 대신 O(log N) 달성 가능",
        "이진 탐색(Binary Search): lo=0, hi=n-1로 초기화.\n"
        "mid=(lo+hi)/2 계산 후 arr[mid]==target이면 반환.\n"
        "arr[mid]<target이면 lo=mid+1, 크면 hi=mid-1.\n"
        "lo>hi가 되면 -1 반환.",
        # C++ 코드
        "int binarySearch(const vector<int>& arr, int target) {\n"
        "    int lo = 0, hi = (int)arr.size() - 1;\n"
        "    while (lo <= hi) {\n"
        "        int mid = lo + (hi - lo) / 2;  // 오버플로 방지\n"
        "        if (arr[mid] == target) return mid;\n"
        "        if (arr[mid] < target)  lo = mid + 1;\n"
        "        else                    hi = mid - 1;\n"
        "    }\n"
        "    return -1;\n"
        "}",
        # C# 코드
        "int BinarySearch(int[] arr, int target) {\n"
        "    int lo = 0, hi = arr.Length - 1;\n"
        "    while (lo <= hi) {\n"
        "        int mid = lo + (hi - lo) / 2;\n"
        "        if (arr[mid] == target) return mid;\n"
        "        if (arr[mid] < target)  lo = mid + 1;\n"
        "        else                    hi = mid - 1;\n"
        "    }\n"
        "    return -1;\n"
        "}",
        "O(log N) 시간, O(1) 공간",
        "Easy",
    ),
    (
        "주어진 N개 수를 빠르게 정렬하세요 (비교 기반 최적).",
        "일반 비교 정렬 / 평균 최적화 우선\n추가 메모리 O(log N) 허용\n불안정 정렬 허용 가능",
        "비교 기반 정렬 하한 O(N log N) — 이를 평균적으로 달성하면서 캐시 효율이 좋음\n"
        "분할 정복으로 부분 배열 독립 처리 → 병렬화 용이",
        "퀵소트(Quick Sort): 피벗 선택(랜덤 또는 중앙값) 후 파티션.\n"
        "피벗보다 작은 요소는 왼쪽, 큰 요소는 오른쪽으로 이동.\n"
        "각 부분을 재귀 호출. 최악 O(N²)이지만 랜덤 피벗 시 평균 O(N log N).",
        # C++
        "void quickSort(vector<int>& a, int l, int r) {\n"
        "    if (l >= r) return;\n"
        "    // 랜덤 피벗으로 최악 케이스 방지\n"
        "    swap(a[l + rand() % (r - l + 1)], a[r]);\n"
        "    int pivot = a[r], i = l - 1;\n"
        "    for (int j = l; j < r; ++j)\n"
        "        if (a[j] <= pivot) swap(a[++i], a[j]);\n"
        "    swap(a[++i], a[r]);\n"
        "    quickSort(a, l, i - 1);\n"
        "    quickSort(a, i + 1, r);\n"
        "}",
        # C#
        "void QuickSort(int[] a, int l, int r) {\n"
        "    if (l >= r) return;\n"
        "    var rng = new Random();\n"
        "    int pi = l + rng.Next(r - l + 1);\n"
        "    (a[pi], a[r]) = (a[r], a[pi]);\n"
        "    int pivot = a[r], i = l - 1;\n"
        "    for (int j = l; j < r; j++)\n"
        "        if (a[j] <= pivot) { i++; (a[i], a[j]) = (a[j], a[i]); }\n"
        "    i++; (a[i], a[r]) = (a[r], a[i]);\n"
        "    QuickSort(a, l, i - 1);\n"
        "    QuickSort(a, i + 1, r);\n"
        "}",
        "평균 O(N log N) / 최악 O(N²), 공간 O(log N) 스택",
        "Normal",
    ),
    (
        "그래프의 두 노드 사이 최단 거리를 구하세요 (가중치 ≥ 0).",
        "가중치 비음수\n단일 출발점\n인접 리스트 그래프",
        "음수 없음 → Dijkstra 적용 가능\n우선순위 큐(min-heap)로 현재 최솟값 노드 빠르게 선택\n"
        "Bellman-Ford(O(VE))보다 O((V+E) log V)로 효율적",
        "Dijkstra: dist[src]=0, 나머지 INF 초기화.\n"
        "min-heap에 (0, src) 삽입. 팝 시 현재 거리가 기록보다 크면 스킵.\n"
        "인접 노드 완화: dist[v] > dist[u]+w이면 갱신 후 큐 삽입.",
        # C++
        "vector<int> dijkstra(int src, int n,\n"
        "    const vector<vector<pair<int,int>>>& adj) {\n"
        "    vector<int> dist(n, INT_MAX);\n"
        "    priority_queue<pair<int,int>,\n"
        "        vector<pair<int,int>>, greater<>> pq;\n"
        "    dist[src] = 0; pq.push({0, src});\n"
        "    while (!pq.empty()) {\n"
        "        auto [d, u] = pq.top(); pq.pop();\n"
        "        if (d > dist[u]) continue;  // 구버전 스킵\n"
        "        for (auto [v, w] : adj[u])\n"
        "            if (dist[u] + w < dist[v]) {\n"
        "                dist[v] = dist[u] + w;\n"
        "                pq.push({dist[v], v});\n"
        "            }\n"
        "    }\n"
        "    return dist;\n"
        "}",
        # C#
        "int[] Dijkstra(int src, int n,\n"
        "    List<(int v, int w)>[] adj) {\n"
        "    var dist = new int[n];\n"
        "    Array.Fill(dist, int.MaxValue);\n"
        "    dist[src] = 0;\n"
        "    // (dist, node) 최소 힙\n"
        "    var pq = new SortedSet<(int d, int u)>()\n"
        "        { (0, src) };\n"
        "    while (pq.Count > 0) {\n"
        "        var (d, u) = pq.Min; pq.Remove(pq.Min);\n"
        "        if (d > dist[u]) continue;\n"
        "        foreach (var (v, w) in adj[u])\n"
        "            if (dist[u] + w < dist[v]) {\n"
        "                pq.Remove((dist[v], v));\n"
        "                dist[v] = dist[u] + w;\n"
        "                pq.Add((dist[v], v));\n"
        "            }\n"
        "    }\n"
        "    return dist;\n"
        "}",
        "O((V+E) log V) 시간, O(V) 공간",
        "Hard",
    ),
    (
        "N개 노드의 그래프에서 최소 신장 트리(MST) 비용을 구하세요.",
        "간선 가중치 있음\n연결 그래프\n사이클 없는 트리 N-1 간선",
        "간선 중심 → Kruskal이 간단\n"
        "Union-Find로 사이클 감지 O(α(N)) ≈ O(1)\n"
        "Prim은 밀집 그래프에서 유리",
        "Kruskal: 간선을 가중치 오름차순 정렬.\n"
        "Union-Find 초기화. 각 간선(u,v,w)을 순서대로 확인,\n"
        "find(u)≠find(v)이면 선택 후 union. N-1개 선택 완료 시 종료.",
        # C++
        "struct Edge { int u, v, w; };\n"
        "int parent[100001];\n"
        "int find(int x) {\n"
        "    return parent[x] == x ? x\n"
        "        : parent[x] = find(parent[x]); // 경로 압축\n"
        "}\n"
        "long long kruskal(int n, vector<Edge>& edges) {\n"
        "    iota(parent, parent + n + 1, 0);\n"
        "    sort(edges.begin(), edges.end(),\n"
        "        [](auto& a, auto& b){ return a.w < b.w; });\n"
        "    long long cost = 0; int cnt = 0;\n"
        "    for (auto& e : edges) {\n"
        "        int pu = find(e.u), pv = find(e.v);\n"
        "        if (pu == pv) continue;\n"
        "        parent[pu] = pv;\n"
        "        cost += e.w;\n"
        "        if (++cnt == n - 1) break;\n"
        "    }\n"
        "    return cost;\n"
        "}",
        # C#
        "record Edge(int U, int V, int W);\n"
        "int[] parent;\n"
        "int Find(int x) =>\n"
        "    parent[x] == x ? x : parent[x] = Find(parent[x]);\n"
        "long Kruskal(int n, List<Edge> edges) {\n"
        "    parent = Enumerable.Range(0, n+1).ToArray();\n"
        "    edges.Sort((a, b) => a.W - b.W);\n"
        "    long cost = 0; int cnt = 0;\n"
        "    foreach (var e in edges) {\n"
        "        int pu = Find(e.U), pv = Find(e.V);\n"
        "        if (pu == pv) continue;\n"
        "        parent[pu] = pv; cost += e.W;\n"
        "        if (++cnt == n - 1) break;\n"
        "    }\n"
        "    return cost;\n"
        "}",
        "O(E log E) 정렬, O(α(V)) Union-Find",
        "Hard",
    ),
    (
        "배열에서 부분 합이 S 이상이 되는 가장 짧은 연속 부분 배열 길이를 구하세요.",
        "모든 원소 양수\n연속 부분 배열\n합 ≥ S인 최소 길이",
        "양수 보장 → 오른쪽 포인터 이동 시 합이 단조 증가\n"
        "슬라이딩 윈도우(투 포인터)로 O(N) 달성 가능\n정렬·이진탐색 O(N log N) 대비 우수",
        "투 포인터: lo=0, 현재합=0.\n"
        "hi를 오른쪽으로 이동하며 현재합 += arr[hi].\n"
        "현재합 ≥ S이면 ans = min(ans, hi-lo+1), 현재합 -= arr[lo], lo++.\n"
        "hi가 끝에 도달할 때까지 반복.",
        # C++
        "int minSubArray(vector<int>& arr, int s) {\n"
        "    int lo = 0, cur = 0, ans = INT_MAX;\n"
        "    for (int hi = 0; hi < (int)arr.size(); ++hi) {\n"
        "        cur += arr[hi];\n"
        "        while (cur >= s) {\n"
        "            ans = min(ans, hi - lo + 1);\n"
        "            cur -= arr[lo++];\n"
        "        }\n"
        "    }\n"
        "    return ans == INT_MAX ? 0 : ans;\n"
        "}",
        # C#
        "int MinSubArray(int[] arr, int s) {\n"
        "    int lo = 0, cur = 0, ans = int.MaxValue;\n"
        "    for (int hi = 0; hi < arr.Length; hi++) {\n"
        "        cur += arr[hi];\n"
        "        while (cur >= s) {\n"
        "            ans = Math.Min(ans, hi - lo + 1);\n"
        "            cur -= arr[lo++];\n"
        "        }\n"
        "    }\n"
        "    return ans == int.MaxValue ? 0 : ans;\n"
        "}",
        "O(N) 시간, O(1) 공간",
        "Normal",
    ),
    (
        "주어진 배열에서 가장 긴 증가하는 부분 수열(LIS)의 길이를 구하세요.",
        "순서 유지(연속 불필요)\n오름차순 증가\n최대 길이 반환",
        "DP O(N²)도 가능하나 이분탐색+그리디로 O(N log N) 달성\n"
        "patience sorting 원리: 덱에 최솟값 유지로 길이 최적화",
        "그리디+이분탐색: tails 배열 유지.\n"
        "각 원소 x에 대해 tails에서 x 이상인 첫 위치를 이분탐색으로 찾아 교체.\n"
        "끝을 벗어나면 append. tails의 크기가 LIS 길이.",
        # C++
        "int lis(vector<int>& nums) {\n"
        "    vector<int> tails;\n"
        "    for (int x : nums) {\n"
        "        auto it = lower_bound(tails.begin(), tails.end(), x);\n"
        "        if (it == tails.end()) tails.push_back(x);\n"
        "        else                   *it = x;\n"
        "    }\n"
        "    return (int)tails.size();\n"
        "}",
        # C#
        "int LIS(int[] nums) {\n"
        "    var tails = new List<int>();\n"
        "    foreach (int x in nums) {\n"
        "        int idx = tails.BinarySearch(x);\n"
        "        if (idx < 0) idx = ~idx;\n"
        "        if (idx == tails.Count) tails.Add(x);\n"
        "        else                    tails[idx] = x;\n"
        "    }\n"
        "    return tails.Count;\n"
        "}",
        "O(N log N) 시간, O(N) 공간",
        "Hard",
    ),
    (
        "N×M 2D 배열에서 1로 연결된 섬의 수를 구하세요 (DFS/BFS).",
        "상하좌우 4방향 연결\n방문 처리 필수\n경계 조건 체크",
        "그래프 연결 요소 탐색 → DFS/BFS 자연스럽게 적용\n"
        "Union-Find도 가능하나 DFS가 구현이 직관적\n"
        "BFS는 큐 메모리 사용하나 스택 오버플로 걱정 없음",
        "DFS 접근: 전체 셀 순회.\n"
        "cell[r][c]==1이고 미방문이면 DFS 시작 후 카운트 증가.\n"
        "DFS 내부: 현재 셀 방문 처리, 4방향으로 재귀.",
        # C++
        "int numIslands(vector<vector<char>>& grid) {\n"
        "    int R = grid.size(), C = grid[0].size(), cnt = 0;\n"
        "    function<void(int,int)> dfs = [&](int r, int c) {\n"
        "        if (r<0||r>=R||c<0||c>=C||grid[r][c]!='1') return;\n"
        "        grid[r][c] = '0'; // 방문 표시\n"
        "        dfs(r+1,c); dfs(r-1,c);\n"
        "        dfs(r,c+1); dfs(r,c-1);\n"
        "    };\n"
        "    for (int r = 0; r < R; ++r)\n"
        "        for (int c = 0; c < C; ++c)\n"
        "            if (grid[r][c] == '1') { dfs(r, c); ++cnt; }\n"
        "    return cnt;\n"
        "}",
        # C#
        "int NumIslands(char[][] grid) {\n"
        "    int R=grid.Length, C=grid[0].Length, cnt=0;\n"
        "    void Dfs(int r, int c) {\n"
        "        if (r<0||r>=R||c<0||c>=C||grid[r][c]!='1') return;\n"
        "        grid[r][c]='0';\n"
        "        Dfs(r+1,c); Dfs(r-1,c);\n"
        "        Dfs(r,c+1); Dfs(r,c-1);\n"
        "    }\n"
        "    for (int r=0;r<R;r++)\n"
        "        for (int c=0;c<C;c++)\n"
        "            if (grid[r][c]=='1') { Dfs(r,c); cnt++; }\n"
        "    return cnt;\n"
        "}",
        "O(N×M) 시간·공간",
        "Normal",
    ),
    (
        "배열에서 중복 없이 모든 부분집합을 출력하세요 (백트래킹).",
        "원소 중복 없음 (또는 정렬 후 중복 제거)\n순서 무관한 부분집합\n2^N 가짓수",
        "상태 공간 트리: 각 원소 선택/미선택 분기\n"
        "백트래킹으로 불필요한 분기 조기 종료 가능\n"
        "비트마스크 O(2^N×N)도 가능하나 N≤20이면 허용",
        "DFS 백트래킹: start 인덱스부터 순차 선택.\n"
        "매 호출마다 현재 조합을 결과에 추가.\n"
        "i=start~n-1 반복: 선택 추가 → 재귀 → 선택 제거(백트래킹).",
        # C++
        "void subsets(vector<int>& nums, int start,\n"
        "    vector<int>& cur, vector<vector<int>>& res) {\n"
        "    res.push_back(cur);\n"
        "    for (int i = start; i < (int)nums.size(); ++i) {\n"
        "        cur.push_back(nums[i]);\n"
        "        subsets(nums, i + 1, cur, res);\n"
        "        cur.pop_back();\n"
        "    }\n"
        "}",
        # C#
        "void Subsets(int[] nums, int start,\n"
        "    List<int> cur, List<List<int>> res) {\n"
        "    res.Add(new List<int>(cur));\n"
        "    for (int i = start; i < nums.Length; i++) {\n"
        "        cur.Add(nums[i]);\n"
        "        Subsets(nums, i+1, cur, res);\n"
        "        cur.RemoveAt(cur.Count - 1);\n"
        "    }\n"
        "}",
        "O(2^N × N) 시간, O(N) 재귀 스택",
        "Normal",
    ),
    (
        "N×N 격자에서 좌상단(0,0)에서 우하단(N-1,N-1)까지의 경로 수를 구하세요.",
        "이동: 오른쪽 또는 아래로만\n장애물 없음 (또는 있음)\n나머지 연산 여부 확인",
        "최적 부분 구조: dp[r][c] = dp[r-1][c] + dp[r][c-1]\n"
        "중복 부분 문제 → 메모이제이션 또는 바텀업 DP\n수학적으로 C(2N-2, N-1)이지만 장애물 있으면 DP 필수",
        "바텀업 DP: dp[0][0]=1 초기화.\n"
        "첫 행/열은 1로 채움. 나머지 dp[r][c]=dp[r-1][c]+dp[r][c-1].",
        # C++
        "long long uniquePaths(int n) {\n"
        "    vector<vector<long long>> dp(n, vector<long long>(n, 1));\n"
        "    for (int r = 1; r < n; ++r)\n"
        "        for (int c = 1; c < n; ++c)\n"
        "            dp[r][c] = dp[r-1][c] + dp[r][c-1];\n"
        "    return dp[n-1][n-1];\n"
        "}",
        # C#
        "long UniquePaths(int n) {\n"
        "    long[,] dp = new long[n, n];\n"
        "    for (int r=0;r<n;r++) dp[r,0]=1;\n"
        "    for (int c=0;c<n;c++) dp[0,c]=1;\n"
        "    for (int r=1;r<n;r++)\n"
        "        for (int c=1;c<n;c++)\n"
        "            dp[r,c]=dp[r-1,c]+dp[r,c-1];\n"
        "    return dp[n-1,n-1];\n"
        "}",
        "O(N²) 시간·공간 (공간 O(N)으로 최적화 가능)",
        "Normal",
    ),
    (
        "종이를 가로 또는 세로로 여러 번 반으로 접은 뒤, 일부 구역을 잘라낸 후 펼쳤을 때 결과를 출력하세요.",
        "접기마다 서로 맞닿는 좌표 쌍 연결\n잘라낼 때 대표 원소에만 cut 플래그\n접기 순서 보존 필수",
        "접힘으로 생긴 등가 관계 → Union-Find로 O(α) 처리\n"
        "잘라낼 때 대표 노드 1개만 마킹 → 전파 불필요\n"
        "가로접기: r↔(h-1-r), 세로접기: c↔(w-1-c)",
        "1. 접기 연산마다 대응 좌표 쌍 union.\n"
        "2. 잘라낼 구역의 대표 노드 cut=true.\n"
        "3. 펼치기: 모든 (r,c)에서 find()의 cut 값으로 0/1 출력.",
        # C++
        "// Union-Find + cut 플래그 (핵심 구조)\n"
        "struct UF {\n"
        "    vector<int> p, cut;\n"
        "    UF(int n): p(n), cut(n,0) { iota(p.begin(),p.end(),0); }\n"
        "    int find(int x){\n"
        "        return p[x]==x?x:p[x]=find(p[x]);\n"
        "    }\n"
        "    void unite(int a,int b){\n"
        "        a=find(a); b=find(b);\n"
        "        if(a!=b) p[a]=b;\n"
        "    }\n"
        "    void setCut(int x){ cut[find(x)]=1; }\n"
        "    bool isCut(int x){ return cut[find(x)]; }\n"
        "};",
        # C#
        "class UF {\n"
        "    int[] p, cut;\n"
        "    public UF(int n) {\n"
        "        p=Enumerable.Range(0,n).ToArray();\n"
        "        cut=new int[n];\n"
        "    }\n"
        "    public int Find(int x) =>\n"
        "        p[x]==x?x:p[x]=Find(p[x]);\n"
        "    public void Unite(int a,int b){\n"
        "        a=Find(a);b=Find(b);\n"
        "        if(a!=b)p[a]=b;\n"
        "    }\n"
        "    public void SetCut(int x){ cut[Find(x)]=1; }\n"
        "    public bool IsCut(int x){ return cut[Find(x)]==1; }\n"
        "}",
        "O(N×M×α) 접기, O(N×M) 출력",
        "Hard",
    ),
]

HEADERS = ["#", "질문", "핵심 조건", "선택 이유", "모범 답변",
           "C++ 코드", "C# 코드", "시간복잡도", "난이도"]
COL_W   = [4,   40,   28,      28,       42,
           52,       52,        16,        10]


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "알고리즘"

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
    ROW_FILLS   = [
        PatternFill("solid", fgColor="D9E1F2"),
        PatternFill("solid", fgColor="EEF2FB"),
    ]
    ROW_FONT  = Font(name="Consolas", size=9)
    META_FONT = Font(name="맑은 고딕", size=9)

    # 헤더
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # 데이터
    for ri, card in enumerate(ALGO_CARDS, 1):
        fill = ROW_FILLS[ri % 2]
        vals = [ri, *card]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri + 1, column=ci, value=str(v) if v else "")
            c.fill   = fill
            c.border = BORDER
            # 코드 컬럼은 Consolas, 나머지 맑은 고딕
            c.font = ROW_FONT if ci in (6, 7) else META_FONT
            c.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if ci in (1, 9) else "left",
            )
        ws.row_dimensions[ri + 1].height = 80

    for ci, w in enumerate(COL_W, 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w

    ws.freeze_panes = "B2"

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[생성] {OUT_PATH}  ({len(ALGO_CARDS)}개 카드)")
    print("→ DB 삽입: python scripts/import_cards.py --db")


if __name__ == "__main__":
    create_template()
