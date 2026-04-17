"""
create_company_template.py  —  기업 기출 카드 템플릿 생성기
=============================================================
card/company_interview_qa.xlsx 를 샘플 데이터와 함께 생성합니다.
데이터를 추가·수정한 뒤 import_cards.py --db 로 DB에 삽입하세요.

컬럼:  # | 회사명 | 질문 | 모범 답변 | 난이도(Easy/Normal/Hard)

사용:
  python scripts/create_company_template.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = os.path.join(os.path.dirname(__file__), "..")
OUT_PATH = os.path.join(BASE, "card", "company_interview_qa.xlsx")

THIN   = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SAMPLE_CARDS = [
    # (회사명, 질문, 모범 답변, 난이도)
    (
        "넥슨",
        "게임 서버의 동시 접속자 수를 늘리기 위한 스케일아웃 전략을 설명하세요.",
        "수평 확장(Scale-Out)으로 서버 인스턴스를 추가하고, 로드 밸런서(L4/L7)로 트래픽을 분산합니다. "
        "세션 상태는 Redis 같은 외부 저장소에 두어 어느 인스턴스에서도 처리 가능하게 합니다. "
        "DB는 읽기 복제본(Read Replica)을 두어 조회 부하를 분리하고, "
        "핫스팟이 생기는 테이블은 샤딩(Sharding)을 적용합니다.",
        "Hard",
    ),
    (
        "넥슨",
        "UE5 Lumen과 Nanite를 각각 언제 사용하면 효과적인가요?",
        "Nanite는 폴리곤 수가 매우 많은 스태틱 메시(건물·지형·소품)에 적합하며, "
        "런타임에 자동 LOD를 처리해 줘 아티스트가 LOD를 수동으로 만들 필요가 없습니다. "
        "Lumen은 동적 글로벌 일루미네이션으로, 씬 라이팅이 실시간으로 변하는 환경(낮/밤 사이클, 폭발 등)에 유효합니다. "
        "단, 모바일·저사양 기기에서는 두 기능 모두 비용이 크므로 PC·콘솔 타깃 프로젝트에 적용하는 것이 일반적입니다.",
        "Normal",
    ),
    (
        "크래프톤",
        "오픈 월드 게임에서 World Partition(레벨 스트리밍)을 사용할 때 주의할 점은?",
        "셀(Cell) 크기를 너무 작게 설정하면 스트리밍 빈도가 높아져 I/O 병목이 생깁니다. "
        "Data Layers로 낮/밤·이벤트 오브젝트를 분리해 필요할 때만 로드합니다. "
        "액터 배치 시 셀 경계를 넘나드는 오브젝트(대형 건물 등)는 Always Loaded로 설정하거나 "
        "HLOD(Hierarchical LOD)를 활용해 원거리 렌더링 비용을 줄입니다.",
        "Hard",
    ),
    (
        "크래프톤",
        "배틀로얄 게임에서 플레이어 간 충돌 판정을 서버에서 처리하는 이유는?",
        "클라이언트 권위(Client Authority) 방식은 치트에 취약합니다. "
        "서버 권위(Server Authority) 모델에서는 클라이언트가 입력(이동·사격)만 전송하고, "
        "서버가 히트 판정·위치 보정을 수행한 뒤 결과를 브로드캐스트합니다. "
        "레이턴시 보상(Lag Compensation)을 위해 서버는 과거 N프레임의 플레이어 위치를 기록해 두고, "
        "클라이언트 타임스탬프 기준으로 히트 판정을 재연산합니다.",
        "Hard",
    ),
    (
        "넷마블",
        "모바일 게임에서 배터리 소모와 발열을 줄이기 위한 렌더링 최적화 방법은?",
        "1) 타깃 프레임을 30fps로 고정(UE5의 t.MaxFPS)해 GPU 과부하를 방지합니다. "
        "2) Mobile HDR를 끄고 Forward Shading을 사용해 렌더 패스를 줄입니다. "
        "3) 동적 그림자 대신 사전 계산된 라이트맵(Precomputed Lighting)을 활용합니다. "
        "4) 텍스처는 ASTC 압축, Mip Map을 적용해 메모리 대역폭을 절감합니다. "
        "5) DrawCall 수를 줄이기 위해 Instanced Static Mesh와 Merged Actor를 활용합니다.",
        "Normal",
    ),
    (
        "넷마블",
        "C++에서 캐시 지역성(Cache Locality)이 게임 성능에 미치는 영향과 개선 방법은?",
        "CPU 캐시 미스가 발생하면 메인 메모리 접근으로 수백 사이클 대기가 생깁니다. "
        "게임 오브젝트를 포인터 배열 대신 값 타입 배열(SoA 패턴)로 배치하면 "
        "연속된 메모리를 읽으므로 캐시 히트율이 높아집니다. "
        "UE5에서는 TArray<FMyStruct> 대신 개별 TArray<float>를 속성별로 분리하는 방식이나 "
        "Mass Entity(ECS) 컴포넌트를 활용할 수 있습니다.",
        "Hard",
    ),
    (
        "엔씨소프트",
        "MMORPG에서 수천 명의 플레이어가 같은 필드에 있을 때 동기화 부하를 어떻게 줄이나요?",
        "AOI(Area of Interest) 시스템으로 일정 반경 내 플레이어에게만 업데이트를 전송합니다. "
        "이동 패킷을 매 프레임 보내지 않고 임계값 이상 변화가 있을 때만 전송(Delta Compression)합니다. "
        "플레이어 밀집 구역은 채널(인스턴스)로 분리해 서버당 부하를 제한합니다. "
        "원거리 플레이어는 위치 업데이트 주기를 낮춰(Interest Level) 패킷 수를 줄입니다.",
        "Hard",
    ),
    (
        "엔씨소프트",
        "UE5 GameplayAbilitySystem(GAS)을 사용할 때 Attribute와 Effect의 관계를 설명하세요.",
        "AttributeSet에 HP·MP·공격력 같은 수치를 정의합니다. "
        "GameplayEffect는 이 Attribute를 수정하는 데이터 객체로, Instant(즉시)·Duration(지속)·Infinite(영구) 세 타입이 있습니다. "
        "Ability가 Effect를 Apply하면 AbilitySystemComponent가 Attribute를 계산(Base·Current Value 분리)하고 "
        "OnAttributeChange 델리게이트로 UI에 통보합니다. "
        "이 구조는 능력치 변경 로직을 Ability에서 분리해 데이터 드리븐으로 만들 수 있다는 장점이 있습니다.",
        "Hard",
    ),
    (
        "스마일게이트",
        "C++에서 RAII 패턴이란 무엇이며 게임 개발에서 어떻게 활용되나요?",
        "RAII(Resource Acquisition Is Initialization)는 자원 획득을 객체 생성 시점에, "
        "해제를 소멸자 시점에 수행하는 C++ 관용구입니다. "
        "게임에서는 파일 핸들·소켓·뮤텍스·렌더 타깃 등을 래퍼 클래스로 감싸 "
        "예외나 조기 리턴이 발생해도 반드시 해제되도록 합니다. "
        "UE5에서는 FRWScopeLock, TSharedPtr 등이 RAII를 구현한 예입니다.",
        "Normal",
    ),
    (
        "스마일게이트",
        "언리얼 엔진에서 Replication과 RPC의 차이를 설명하세요.",
        "Replication은 서버의 Actor/Component 상태(UPROPERTY)를 클라이언트에 자동으로 동기화하는 메커니즘입니다. "
        "RPC(Remote Procedure Call)는 특정 시점에 명시적으로 함수를 원격 호출하는 방식입니다. "
        "Server RPC: 클라이언트 → 서버 호출(UFUNCTION(Server, Reliable)). "
        "Multicast RPC: 서버 → 모든 클라이언트 호출. "
        "Client RPC: 서버 → 특정 클라이언트 호출. "
        "Reliable은 패킷 손실 시 재전송을 보장하지만 비용이 크므로 중요 이벤트에만 사용합니다.",
        "Normal",
    ),
    (
        "펄어비스",
        "오픈 월드 게임에서 나비메시(NavMesh) 런타임 업데이트를 최적화하는 방법은?",
        "NavMesh는 전체를 재계산하지 않고 변경된 영역만 갱신하는 Dynamic Obstacle을 활용합니다. "
        "UE5에서는 NavMesh Bounds Volume을 여러 조각으로 분할해 부분 업데이트 비용을 낮춥니다. "
        "파괴 가능 오브젝트가 많은 씬에서는 InvokerBasedNavMesh를 사용해 "
        "플레이어 주변 반경만 유지하고 원거리는 해제합니다.",
        "Hard",
    ),
    (
        "펄어비스",
        "멀티스레드 환경에서 GameThread와 RenderThread를 분리하는 이유는?",
        "언리얼 엔진은 GameThread(게임 로직)와 RenderThread(렌더링 커맨드 생성)를 분리해 병렬 실행합니다. "
        "GameThread가 다음 프레임 로직을 처리하는 동안 RenderThread가 이전 프레임 커맨드를 GPU에 제출합니다. "
        "이 구조로 CPU 유휴 시간을 줄이고 프레임 레이트를 높입니다. "
        "단, RenderThread에서 게임 데이터에 직접 접근하면 Race Condition이 생기므로 "
        "ENQUEUE_RENDER_COMMAND 매크로를 통해 안전하게 데이터를 전달해야 합니다.",
        "Hard",
    ),
    # ── 펄어비스 가을 인턴십 직무테스트 기출 ─────────────
    (
        "펄어비스",
        "종이접기 시뮬레이션에서 접힌 칸들의 컷(cut) 상태를 효율적으로 공유하는 자료구조와 알고리즘을 설명하세요.",
        "Union-Find(Disjoint Set) 활용. "
        "접기 연산마다 서로 맞닿는 칸 쌍을 union 처리해 동일 운명(컷 여부)을 공유합니다. "
        "잘라낼 때는 대표 원소(find 결과)에만 cut = true 플래그를 세우면, "
        "링크된 모든 칸을 잘린 것으로 간주할 수 있습니다. "
        "펼치기 후 각 (r,c)의 출력은 find(r,c)의 cut 값을 참조합니다. "
        "시간복잡도: 각 접기 단계 O(nm), 경로 압축 시 find는 O(α(nm)). "
        "인덱스 반전 공식: 가로 접기(상하) r→(h-1-r), 세로 접기(좌우) c→(w-1-c). "
        "접기 순서가 결과를 바꾸므로 연산 순서를 반드시 보존해야 합니다.",
        "Hard",
    ),
    (
        "펄어비스",
        "오름수(자릿수가 항상 증가하는 수열)의 k번째 값을 구하는 효율적인 방법은?",
        "핵심 사실: 0~9 중 L개를 오름차순으로 배치하면 오름수가 정확히 C(10,L)개. "
        "전체 오름수 개수 = ΣC(10,L) (L=1..10) = 2^10 - 1 = 1023개. "
        "접근(조합 카운팅): "
        "1) 길이별 C(10,L)의 누적 합으로 k가 속한 길이 L을 결정. "
        "2) 앞자리부터 작은 숫자 후보를 넣어보며 C(남은숫자수, 남은자리수)로 k를 스킵. "
        "3) k가 정확히 1이 되는 시점의 조합을 출력. "
        "k > 1023이면 존재하지 않으므로 -1(또는 문제 요구값) 반환. "
        "DFS 백트래킹도 가능(최대 1023회 탐색).",
        "Hard",
    ),
    (
        "펄어비스",
        "struct 멤버의 메모리 패딩을 최소화하기 위한 배치 전략은?",
        "CPU는 각 타입의 크기에 맞춰 주소를 정렬(alignment)하므로, "
        "작은 타입이 앞에 오면 뒤에 패딩이 삽입되어 구조체 크기가 커집니다. "
        "전략: 정렬 단위가 큰 타입부터 작은 타입 순으로 배치합니다. "
        "예: double(8B) → int(4B) → short(2B) → char(1B). "
        "반대 배치 시 패딩으로 크기가 최대 2배 이상 불어날 수 있습니다. "
        "컴파일 타임 검증: static_assert(sizeof(S) == N, \"패딩 확인\").",
        "Normal",
    ),
    (
        "펄어비스",
        "IEEE 754 부동소수점 값을 == 로 비교하면 안 되는 이유와 올바른 비교법은?",
        "이진수로 정확히 표현할 수 없는 실수가 존재해 연산마다 오차가 누적됩니다. "
        "예: 0.1 + 0.2 == 0.3 → false (실제 결과 0.30000000000000004). "
        "덧셈 결합법칙도 성립하지 않아 (a+b)+c ≠ a+(b+c)가 될 수 있습니다. "
        "올바른 비교: fabs(a - b) < ε (ε은 허용 오차, 보통 1e-6 ~ 1e-9). "
        "절대 오차와 상대 오차를 함께 고려하면 더 견고합니다.",
        "Normal",
    ),
    (
        "펄어비스",
        "댕글링 포인터(Dangling Pointer)와 널 포인터(Null Pointer)의 차이와 방지 방법은?",
        "댕글링 포인터: delete/free 이후 해당 포인터를 계속 사용(Use-After-Free). "
        "메모리 손상·정의되지 않은 동작(UB)·보안 취약점 유발. "
        "방지: 해제 직후 nullptr 대입, unique_ptr/shared_ptr로 소유권 자동 관리(RAII). "
        "널 포인터: 초기화되지 않았거나 의도적으로 nullptr인 포인터. "
        "역참조 시 런타임 오류(segfault) 발생. "
        "방지: 선언과 동시에 초기화, 사용 전 nullptr 체크(if (p != nullptr)).",
        "Normal",
    ),
    (
        "펄어비스",
        "Kruskal 알고리즘으로 MST를 구성할 때 사용되지 않는 간선을 찾는 방법은?",
        "Kruskal: 간선을 가중치 오름차순 정렬 후 Union-Find로 사이클 없는 간선만 선택. "
        "시간복잡도 O(E log E). "
        "선택되지 않은 간선 = 사이클을 유발하는 간선 = 사용되지 않는 간선. "
        "구하는 방법: 전체 간선 집합 − MST에 포함된 간선 집합. "
        "보완: Prim 알고리즘(O(E log V))은 밀집 그래프에서 유리합니다.",
        "Normal",
    ),
    (
        "펄어비스",
        "해시 충돌 해결 방법 두 가지(체이닝, 개방 주소법)를 설명하세요.",
        "체이닝(Chaining): 각 버킷을 연결 리스트로 구성. "
        "같은 해시값 원소를 체인에 추가. Load Factor가 높으면 탐색이 O(n)에 가까워지므로 리사이즈 필요. "
        "개방 주소법(Open Addressing): 충돌 시 다음 빈 버킷을 탐사. "
        "선형 탐사(linear), 이차 탐사(quadratic), 이중 해싱(double hashing) 방식 존재. "
        "캐시 지역성이 좋지만 클러스터링(군집화) 문제 발생. "
        "일반적으로 Load Factor 0.75 이하로 유지해 성능 보장.",
        "Normal",
    ),
    (
        "펄어비스",
        "멀티스레딩 공유 자원 보호 수단과 데드락 회피 전략을 설명하세요.",
        "보호 수단: mutex(상호 배제), spinlock(짧은 대기), semaphore(자원 개수 제한), "
        "RW lock(읽기 다수·쓰기 소수 상황에 효율적). "
        "데드락 4조건: 상호 배제·점유 대기·비선점·순환 대기. "
        "회피 전략: 1) Lock Ordering — 모든 스레드가 동일 순서로 락 획득. "
        "2) try_lock + 백오프 — 실패 시 이미 획득한 락 해제 후 재시도. "
        "3) 크리티컬 섹션 최소화. "
        "4) atomic 연산·Lock-Free 자료구조로 락 제거. "
        "추가: false sharing 방지를 위해 공유 변수를 캐시라인(64B) 단위로 정렬.",
        "Hard",
    ),
    (
        "펄어비스",
        "스택 오버플로우(Stack Overflow)의 원인과 해결 방법은?",
        "원인: 1) 종료 조건 오류로 인한 무한 재귀. "
        "2) 재귀 깊이가 너무 깊어 스택 한도 초과. "
        "3) 스택에 너무 큰 지역 배열 선언(예: int arr[100000]). "
        "해결: 1) 꼬리 재귀(Tail Recursion) → 반복문으로 전환. "
        "2) 명시적 스택(std::stack, TArray 등 힙 메모리)으로 DFS 대체. "
        "3) 대형 배열은 힙에 동적 할당(new / TArray). "
        "4) 스레드 생성 시 스택 크기 명시적 지정(플랫폼 API).",
        "Normal",
    ),
    (
        "펄어비스",
        "게임 수학에서 벡터 내적(Dot Product)과 외적(Cross Product)의 활용 차이를 설명하세요.",
        "내적: a·b = |a||b|cosθ = axbx + ayby + azbz (스칼라 반환). "
        "활용: 두 벡터 사이 각도 계산, 정사영(Projection), 시야각 판정(전방/후방), "
        "Lambert 조명 강도(법선·광원 방향 내적). "
        "외적(3D 전용): a×b = |a||b|sinθ 방향 벡터(벡터 반환). "
        "활용: 두 벡터에 수직인 법선 계산(폴리곤 노멀), "
        "회전 방향 판정(왼쪽/오른쪽), 평행사변형 넓이, "
        "탄젠트/바이탄젠트 계산(노멀 매핑).",
        "Normal",
    ),
    (
        "카카오게임즈",
        "TCP와 UDP를 게임에서 어떻게 구분해서 사용하나요?",
        "TCP는 신뢰성(재전송·순서 보장)을 제공하지만 레이턴시가 높습니다. "
        "로그인·인벤토리·결제처럼 데이터 유실이 치명적인 경우에 적합합니다. "
        "UDP는 빠르지만 패킷 손실·순서 역전이 발생할 수 있습니다. "
        "실시간 이동 동기화·음성 채팅처럼 최신 상태가 중요하고 "
        "약간의 손실이 허용되는 경우에 사용합니다. "
        "많은 게임에서 UDP 위에 재전송·혼잡 제어를 직접 구현한 RUDP를 씁니다.",
        "Normal",
    ),
    (
        "카카오게임즈",
        "게임 서버에서 데드락(Deadlock)이 발생하는 조건과 예방 방법은?",
        "데드락 발생 4조건: 상호 배제, 점유 대기, 비선점, 순환 대기. "
        "예방법: 1) Lock Ordering — 모든 스레드가 동일한 순서로 락을 획득합니다. "
        "2) Try-Lock + 백오프 — 락 획득 실패 시 이미 획득한 락을 해제하고 재시도합니다. "
        "3) 락 범위 최소화 — 크리티컬 섹션을 짧게 유지해 경합을 줄입니다. "
        "4) Lock-Free 자료구조(atomic 연산) 활용으로 뮤텍스 자체를 제거합니다.",
        "Normal",
    ),
]


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기업기출"

    HEADER_FILL = PatternFill("solid", fgColor="4A235A")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    ROW_FILLS   = [
        PatternFill("solid", fgColor="E8DAEF"),
        PatternFill("solid", fgColor="F5EEF8"),
    ]
    ROW_FONT = Font(name="맑은 고딕", size=10)

    headers = ["#", "회사명", "질문", "모범 답변", "난이도"]
    col_w   = [5,   12,      44,    72,          10]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ri, (company, q, a, diff) in enumerate(SAMPLE_CARDS, 1):
        fill = ROW_FILLS[ri % 2]
        vals = [ri, company, q, a, diff]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri + 1, column=ci, value=v)
            c.fill = fill; c.font = ROW_FONT; c.border = BORDER
            c.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if ci in (1, 2, 5) else "left",
            )
        ws.row_dimensions[ri + 1].height = 60

    for ci, w in enumerate(col_w, 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[생성] {OUT_PATH}  ({len(SAMPLE_CARDS)}개 샘플 포함)")
    print("데이터를 추가·수정한 뒤 아래 명령어로 DB에 삽입하세요:")
    print("  python scripts/import_cards.py --db")


if __name__ == "__main__":
    create_template()
