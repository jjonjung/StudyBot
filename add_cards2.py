import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER = r"C:\Users\EJ\Desktop\Fork\StudyBot\card\master"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)
def font(bold=False, size=11, color="000000", name="맑은 고딕"):
    return Font(bold=bold, size=size, color=color, name=name)
def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NEW_CARDS = {

"01_cpp": [
    (
        "프로젝트에서 맡은 역할과 기여 범위를 어떻게 설명하는가?",
        "①팀 구성과 전체 프로젝트 규모 → ②내가 단독 설계·구현한 모듈 → ③공동 작업 부분과 내 기여 비율 → ④성과(기능 완성·수치·피드백) 순으로 구조화. '나는 ~를 담당했다'보다 '해당 모듈의 설계 결정을 내가 내렸고, 이유는 ~이다'로 주도성을 드러냄.",
        "역할, 기여, 주도성, 구조적 설명",
        "중",
        "팀원과 기술 방향이 달랐을 때 어떻게 조율했는가?"
    ),
    (
        "병목 구간을 어떻게 분석하고 측정했는가?",
        "①프로파일러(Unreal Insights / Superluminal / VS Profiler)로 CPU·GPU 핫스팟 확인 → ②의심 구간에 TRACE_CPUPROFILER_EVENT_SCOPE 또는 std::chrono 타이머 삽입 → ③캐시 미스·메모리 할당·스레드 경합 여부 확인 → ④개선 전·후 수치 비교(프레임 시간, 틱 연산 ms). '느린 것 같아서'가 아니라 측정 데이터 기반으로 설명.",
        "프로파일러, 병목, 측정, Unreal Insights, 핫스팟",
        "중상",
        "CPU 병목과 GPU 병목을 구분하는 기준은?"
    ),
    (
        "C++ 메모리 할당/해제(new·delete·malloc·free)의 차이는?",
        "malloc/free: C 함수, 바이트 단위 할당, 생성자·소멸자 미호출. new/delete: C++ 연산자, 타입 크기 자동 계산, 생성자·소멸자 호출. new[]는 delete[]로 해제해야 함(다른 해제 시 UB). 현대 C++에서는 RAII를 위해 unique_ptr/shared_ptr 선호. placement new는 미리 확보한 메모리에 객체를 생성할 때 사용.",
        "new, delete, malloc, free, RAII, UB, placement new",
        "중",
        "new/delete를 직접 쓰지 말아야 하는 이유는?"
    ),
],

"02_cs": [
    (
        "배열·연결 리스트·맵(해시맵)의 차이와 선택 기준은?",
        "배열: 인덱스 O(1) 접근, 삽입·삭제 O(n), 캐시 친화적. 연결 리스트: 삽입·삭제 O(1)(노드 있을 때), 접근 O(n), 캐시 비친화적. 해시맵: 키 기반 O(1) 평균 탐색·삽입·삭제, 순서 없음, 충돌 오버헤드. 선택 기준: 순차 접근 빈도·삽입/삭제 빈도·키 조회 필요 여부. 게임 서버: 세션 관리는 unordered_map, 최근 로그는 deque, 고정 수 버퍼는 array.",
        "배열, 연결 리스트, 해시맵, 캐시, 선택 기준",
        "중",
        "std::vector와 std::list를 실제로 어떻게 선택하는가?"
    ),
],

"04_unreal": [
    (
        "AI 캐릭터 설계 시 구조, 확장성, 유지보수 관점을 어떻게 고려하는가?",
        "구조: Behavior Tree + Blackboard로 상태·조건 분리, Task/Decorator/Service 단위 재사용. 확장성: 새 상태 추가 시 기존 Task 수정 없이 Composite 하위에 Task 추가. 유지보수: AI 파라미터(시야, 공격 범위)는 DataAsset으로 분리해 에디터에서 조정 가능하게 설계. 실제 InfinityFighter에서 적 AI Sequence에 Patrol→Chase→Attack Task를 추가하며 AI 행동을 확장한 경험.",
        "Behavior Tree, DataAsset, 확장성, 유지보수, Task",
        "중상",
        "AI 유닛이 100개 이상일 때 성능 병목을 어떻게 관리하는가?"
    ),
    (
        "현재 공부 중인 기술과 이유를 어떻게 답하는가?",
        "게임 서버 구조에 관심을 두고 공부 중. 구체적으로: ①C++ 멀티스레딩(락 없는 자료구조, 메모리 오더) ②소켓 네트워크(Winsock2 Raw TCP, Dead Reckoning, 틱 루프) ③DB 연동(MySQL Connector C++, Stored Procedure). InfinityServer 직접 구현하며 이론이 아닌 실제 문제(패킷 손실, 틱 오버런) 해결 경험 쌓는 중. 다음 목표: IOCP 기반 논블로킹 서버로 확장.",
        "서버, 멀티스레딩, 소켓, 공부 방향, IOCP",
        "하",
        "IOCP와 현재 사용 중인 구조의 차이를 설명하라."
    ),
],

}

HEADER_COLORS = {
    "01_cpp":   "1A3A5C",
    "02_cs":    "1A4731",
    "04_unreal":"3B1F5E",
}
Q_COLORS = {
    "01_cpp":   "D6E4F7",
    "02_cs":    "D6EFE0",
    "04_unreal":"E8DEFF",
}
A_COLORS = {
    "01_cpp":   "EAF4FF",
    "02_cs":    "EAF7EE",
    "04_unreal":"F2ECFF",
}

def append_cards(key, cards):
    path = os.path.join(MASTER, f"{key}.xlsx")
    if not os.path.exists(path):
        print(f"[SKIP] {path} 없음")
        return

    wb = load_wb_safe(path)
    ws_flash  = wb.worksheets[0]
    ws_review = wb.worksheets[1] if len(wb.worksheets) > 1 else None

    hdr_fill = fill(HEADER_COLORS.get(key, "2C3E50"))
    q_fill   = fill(Q_COLORS.get(key,   "F2F2F2"))
    a_fill   = fill(A_COLORS.get(key,   "FAFAFA"))

    existing_qs = set()
    for row in ws_flash.iter_rows(min_row=2, values_only=True):
        if row[1]:
            existing_qs.add(str(row[1]).strip())

    start_no = sum(1 for r in ws_flash.iter_rows(min_row=2, values_only=True) if r[0] is not None) + 1
    added = 0

    for card in cards:
        q, a, kw, diff, tail = card
        if q.strip() in existing_qs:
            continue
        no = start_no + added
        a_len = len(a)
        rh = 36 if a_len < 80 else (56 if a_len < 200 else 80)

        row_q = [no, q, "", kw, diff, tail]
        row_a = ["", "", a, "", "", ""]

        for ri, (row_data, bg, bold_col) in enumerate([(row_q, q_fill, True), (row_a, a_fill, False)]):
            ws_flash.append(row_data)
            cur = ws_flash.max_row
            ws_flash.row_dimensions[cur].height = rh
            for ci, val in enumerate(row_data, 1):
                cell = ws_flash.cell(cur, ci)
                cell.fill = bg
                cell.border = thin_border()
                cell.alignment = align()
                cell.font = font(bold=(ci == 2 and bold_col))

        if ws_review:
            ws_review.append([no, q, a[:120] + ("…" if len(a) > 120 else "")])
            cur = ws_review.max_row
            ws_review.row_dimensions[cur].height = 36
            for ci in range(1, 4):
                cell = ws_review.cell(cur, ci)
                cell.fill = q_fill if ci < 3 else a_fill
                cell.border = thin_border()
                cell.alignment = align()

        existing_qs.add(q.strip())
        added += 1

    wb.save(path)
    total = sum(1 for r in wb.worksheets[0].iter_rows(min_row=2, values_only=True) if r[0] is not None)
    print(f"[OK] {key}.xlsx  +{added}개 추가 → 총 {total // 2}개")

def load_wb_safe(path):
    import openpyxl
    return openpyxl.load_workbook(path)

if __name__ == "__main__":
    for key, cards in NEW_CARDS.items():
        append_cards(key, cards)
    print("\n추가 완료.")
