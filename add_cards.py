import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER = r"C:\Users\EJ\Desktop\Fork\StudyBot\card\master"

# ─── 스타일 헬퍼 ─────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)
def font(bold=False, size=11, color="000000", name="맑은 고딕"):
    return Font(bold=bold, size=size, color=color, name=name)
def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── 추가 카드 데이터 ─────────────────────────────────────────────
# 형식: (질문, 핵심답변, 키워드, 난이도, 꼬리질문)

NEW_CARDS = {

# ════════════════════════════════════════════════════════════════════
# 01_cpp.xlsx — 포인터·레퍼런스·캐스팅·프로젝트 구조 설명
# ════════════════════════════════════════════════════════════════════
"01_cpp": [

    # ── 1.1 프로젝트 설명 구조 ─────────────────────────────────────
    (
        "프로젝트를 면접관에게 설명하는 구조적 순서는?",
        "① 프로젝트 목적·규모\n"
        "② 본인 역할·기여 범위\n"
        "③ 핵심 문제 상황 (현상)\n"
        "④ 원인 분석 (왜 발생했는가)\n"
        "⑤ 해결 전략 선택 이유 (다른 방법과 비교)\n"
        "⑥ 결과 수치 (성능·안정성 등)\n"
        "⑦ 개선 가능 포인트\n\n"
        "핵심: '무엇을 했다'보다 '왜 그렇게 했는지' 중심으로 설명.",
        "PAAR·문제정의·원인분석·결과수치",
        "상",
        "가장 어려웠던 기술적 문제는 무엇이었고, 처음 시도한 방법이 실패했다면 어떻게 방향을 바꿨는가?"
    ),
    (
        "다시 만든다면 개선하고 싶은 점을 묻는 질문에 어떻게 답하는가?",
        "구체적 기술 부채 + 이유 + 개선 방향 3단 구조로 답변.\n\n"
        "예시 (InfinityServer):\n"
        "'스레드-per-클라이언트 모델을 썼는데, 동접자 수가 늘면 스레드 생성 비용과 컨텍스트 스위칭 비용이 선형 증가합니다. "
        "다시 만든다면 IOCP(I/O Completion Port) 기반 비동기 I/O로 전환해 고정 스레드풀에서 이벤트를 처리하는 구조로 바꾸겠습니다. "
        "이를 통해 10,000 동접 기준 스레드 수를 수천 개에서 수십 개로 줄일 수 있습니다.'\n\n"
        "포인트: 단점 인정 → 기술적 근거 → 대안 제시. 자기 인식을 보여주는 질문.",
        "기술부채·IOCP·자기회고·개선방향",
        "상",
        "해당 개선을 실제로 적용하려면 어디서부터 시작하겠는가?"
    ),
    (
        "최적화 경험을 설명할 때 '측정 기준 + 개선 수치 + 선택 이유'를 포함해야 하는 이유는?",
        "'최적화했다'는 주관적 주장에 불과함. 면접관은 아래 3가지를 듣고 싶어 함:\n\n"
        "① 측정 기준: 어떤 도구로 병목을 발견했는가? (Unreal Profiler, Visual Studio Profiler, 로그 타이머 등)\n"
        "② 개선 수치: CPU 40% 감소 / 프레임 55→90fps / 메모리 200MB→80MB 등 수치\n"
        "③ 선택 이유: 왜 다른 방법이 아닌 이 방법을 택했는가? (트레이드오프)\n\n"
        "수치 없는 최적화 설명 = '열심히 했다'와 같은 말. 숫자가 설득력.",
        "병목측정·수치근거·트레이드오프·Profiler",
        "중",
        "측정 없이 감으로 최적화를 시작했다가 실패한 경험이 있는가?"
    ),

    # ── 1.3 포인터 vs 레퍼런스 ──────────────────────────────────────
    (
        "포인터와 레퍼런스의 차이와 각각을 선택하는 기준은?",
        "포인터(T*):\n"
        "- nullptr 가능, 재할당 가능, 배열·동적 할당에 사용\n"
        "- 역참조(*) 필요, 소유권 표현 불명확\n\n"
        "레퍼런스(T&):\n"
        "- nullptr 불가, 초기화 후 재바인딩 불가, 항상 유효한 객체 보장\n"
        "- 함수 인수로 '수정 가능한 객체를 가리킬 때' 선호\n\n"
        "선택 기준:\n"
        "- 선택적(null 가능)이거나 재할당이 필요하면 → 포인터\n"
        "- 항상 유효한 객체, 함수 파라미터 → 레퍼런스\n"
        "- 소유권 명시 → unique_ptr / shared_ptr\n\n"
        "UE5에서는 UObject는 TObjectPtr, 비-UObject는 & 또는 raw pointer 사용.",
        "포인터·레퍼런스·nullptr·소유권·TObjectPtr",
        "중",
        "const T& vs T* const의 차이는? 각각 어디에 쓰는가?"
    ),
    (
        "C++ 캐스팅 4종(static_cast, dynamic_cast, reinterpret_cast, const_cast)의 차이와 위험은?",
        "static_cast:\n"
        "- 컴파일 타임 타입 변환. 관련 있는 타입 간(int→float, 기반→파생).\n"
        "- 안전: 컴파일러가 타입 관계를 확인. 잘못 쓰면 UB.\n\n"
        "dynamic_cast:\n"
        "- 런타임 RTTI 기반. 기반→파생 다운캐스트 안전 확인.\n"
        "- 포인터: 실패 시 nullptr 반환. 레퍼런스: 실패 시 bad_cast 예외.\n"
        "- virtual 함수가 1개 이상 있어야 사용 가능.\n\n"
        "reinterpret_cast:\n"
        "- 비트 수준 강제 재해석. 포인터↔정수 등.\n"
        "- 가장 위험. 거의 사용 금지 (직렬화·하드웨어 레지스터 접근 등 예외적 상황만).\n\n"
        "const_cast:\n"
        "- const/volatile 제거. 원본이 const인 객체 수정 시 UB.\n"
        "- 레거시 API 연동 시 최소한으로 사용.\n\n"
        "게임 코드 원칙: dynamic_cast 남발 → 설계 냄새. 가상 함수/인터페이스로 대체.",
        "static_cast·dynamic_cast·reinterpret_cast·const_cast·RTTI·UB",
        "중",
        "dynamic_cast가 실패하는 경우를 코드로 설명하고, 이를 방지하는 설계 방법은?"
    ),

    # ── 1.7 클래스 메모리 크기 / 다중 상속 ───────────────────────────
    (
        "C++ 클래스 메모리 크기는 어떻게 계산하는가? 패딩과 정렬 규칙은?",
        "기본 규칙:\n"
        "① 각 멤버는 자신의 크기에 맞는 주소에 배치 (alignment)\n"
        "② 구조체 전체 크기는 가장 큰 멤버 크기의 배수로 패딩\n\n"
        "예시:\n"
        "struct A { char c; int i; char d; };\n"
        "  → c(1)+패딩(3)+i(4)+d(1)+패딩(3) = 12바이트\n\n"
        "struct B { int i; char c; char d; };\n"
        "  → i(4)+c(1)+d(1)+패딩(2) = 8바이트\n\n"
        "virtual 함수 있으면 vptr(8바이트) 추가.\n"
        "상속 시 부모 멤버 먼저 배치.\n\n"
        "최적화: 큰 멤버부터 선언해 패딩 최소화. #pragma pack으로 강제 압축 (네트워크 패킷 구조체 등).",
        "sizeof·패딩·alignment·vptr·pragma_pack",
        "중",
        "가상 함수가 있는 클래스와 없는 클래스의 sizeof 차이를 예로 들어라."
    ),
    (
        "다중 상속의 다이아몬드 문제와 virtual 상속으로 해결하는 방법은?",
        "다이아몬드 문제:\n"
        "A → B, A → C, B·C → D 구조에서\n"
        "D가 A의 멤버를 B 경로와 C 경로 두 곳에서 상속 → 중복.\n\n"
        "해결: virtual 상속\n"
        "class B : virtual public A {};\n"
        "class C : virtual public A {};\n"
        "class D : public B, public C {};\n"
        "→ A 인스턴스가 하나만 생성됨 (vptr 추가 비용 발생).\n\n"
        "UE5 원칙: 다중 구현 상속 금지. UInterface(순수 인터페이스)만 다중 상속 허용.\n"
        "→ 다이아몬드 문제를 구조적으로 차단.",
        "다이아몬드·virtual상속·UInterface·다중상속",
        "상",
        "UE5에서 다중 상속 대신 Component 패턴을 쓰는 이유는?"
    ),
    (
        "팩토리 패턴과 컴포지트 패턴을 게임 개발에서 어떻게 활용하는가?",
        "팩토리 패턴:\n"
        "- 객체 생성 로직을 클라이언트에서 분리.\n"
        "- 게임: 적 종류(Goblin/Orc/Dragon)를 스트링·enum으로 생성.\n"
        "- UE5: SpawnActor<T>() + DataAsset 조합이 팩토리 역할.\n"
        "- 장점: 신규 적 추가 시 생성 코드 수정 없음 (OCP).\n\n"
        "컴포지트 패턴:\n"
        "- 단일 객체와 복합 객체를 동일 인터페이스로 처리.\n"
        "- 게임: 스킬 트리 (단일 스킬 / 스킬 그룹 모두 Execute() 호출).\n"
        "- UE5의 Actor-Component 구조가 컴포지트 패턴의 변형.\n\n"
        "선택 기준: '생성 다양화' → 팩토리. '트리 구조 일괄 처리' → 컴포지트.",
        "팩토리·컴포지트·OCP·SpawnActor·DataAsset",
        "중",
        "팩토리 패턴과 추상 팩토리 패턴의 차이는?"
    ),

    # ── 1.6 수학·그래픽스 ─────────────────────────────────────────
    (
        "내적(Dot Product)의 정의와 게임에서 활용 방법은?",
        "정의: A·B = |A||B|cosθ\n"
        "정규화 벡터끼리: A·B = cosθ\n\n"
        "게임 활용:\n"
        "① 시야 판정: 플레이어→적 벡터와 플레이어 전방 벡터의 내적 > cos(시야각/2) → 시야 내\n"
        "② 조명(Lambert): 법선N과 광원방향L의 내적 → 밝기 계산\n"
        "③ 같은 방향 판단: 내적 > 0 → 같은 방향, < 0 → 반대\n"
        "④ 투영: 벡터 A를 B 방향으로 투영한 크기 = A·B̂\n\n"
        "UE5: FVector::DotProduct(A, B) 또는 A | B",
        "DotProduct·cosθ·시야판정·Lambert·투영",
        "중",
        "내적이 0인 벡터 관계는? 게임에서 이 조건을 어디에 쓰는가?"
    ),
    (
        "외적(Cross Product)의 정의와 게임에서 활용 방법은?",
        "정의: A×B = |A||B|sinθ 방향의 수직 벡터 (오른손 법칙)\n\n"
        "게임 활용:\n"
        "① 법선 벡터 계산: 삼각형의 두 엣지 벡터 외적 → 면 법선 (렌더링, 충돌)\n"
        "② 회전 방향 판정: A×B의 Y(혹은 Z) 부호 → 시계/반시계 판단\n"
        "③ 조준 보조: 캐릭터 오른쪽 벡터 = Forward × Up\n"
        "④ 토크 계산: 물리 시뮬레이션에서 회전력\n\n"
        "UE5: FVector::CrossProduct(A, B) 또는 A ^ B\n\n"
        "주의: 외적은 교환 법칙 성립 안 함 (A×B = -(B×A)).",
        "CrossProduct·법선벡터·오른손법칙·회전방향·토크",
        "중",
        "삼각형의 세 꼭짓점이 주어졌을 때 법선 벡터를 구하는 코드를 작성하라."
    ),
    (
        "반사 벡터 공식과 게임에서의 구현 방법은?",
        "공식: R = V - 2(V·N)N\n"
        "V: 입사 벡터(단위벡터), N: 법선 벡터(단위벡터)\n\n"
        "유도: V를 N 방향 성분과 접선 성분으로 분해\n"
        "  V_n = (V·N)N  (N 방향 성분)\n"
        "  V_t = V - V_n  (접선 성분)\n"
        "  R = V_t - V_n = V - 2V_n = V - 2(V·N)N\n\n"
        "게임 활용:\n"
        "① 총알·투사체 리코셰(ricochet): 벽 법선으로 반사 방향 계산\n"
        "② 조명 스페큘러: 반사 벡터와 시선 벡터의 내적 → 하이라이트 강도\n"
        "③ 음향 반사 시뮬레이션\n\n"
        "UE5: FMath::GetReflectionVector(Direction, SurfaceNormal)",
        "반사벡터·R=V-2VNN·리코셰·스페큘러·법선",
        "중",
        "완전 반사(specular)와 난반사(diffuse)의 차이를 벡터 연산 관점에서 설명하라."
    ),
    (
        "좌표계 변환(로컬→월드→뷰→클립 공간)의 흐름을 설명하라.",
        "변환 파이프라인:\n"
        "로컬 공간 → [모델 행렬] → 월드 공간\n"
        "월드 공간 → [뷰 행렬]  → 뷰(카메라) 공간\n"
        "뷰 공간   → [투영 행렬] → 클립 공간\n"
        "클립 공간 → [원근 나누기] → NDC → [뷰포트] → 화면\n\n"
        "각 공간 의미:\n"
        "- 로컬: 메시 원점 기준\n"
        "- 월드: 씬 원점 기준\n"
        "- 뷰: 카메라를 원점으로 변환\n"
        "- 클립: 투영 후 [-w, w] 범위. NDC는 [-1,1]\n\n"
        "UE5에서 자주 쓰는 함수:\n"
        "- GetActorTransform() → 로컬→월드\n"
        "- ProjectWorldLocationToScreen() → 월드→화면\n"
        "- InverseTransformPosition() → 월드→로컬",
        "좌표변환·모델행렬·뷰행렬·투영행렬·NDC·클립공간",
        "상",
        "월드→로컬 변환이 필요한 게임 시나리오를 하나 들어라."
    ),
],

# ════════════════════════════════════════════════════════════════════
# 02_cs.xlsx — 시간복잡도·스택오버플로우·링크드리스트·OS
# ════════════════════════════════════════════════════════════════════
"02_cs": [
    (
        "Big-O 시간복잡도 표기법이란? O(1)·O(log n)·O(n)·O(n log n)·O(n²)를 예와 함께 설명하라.",
        "Big-O: 입력 크기 n이 커질 때 연산 증가율의 상한.\n\n"
        "O(1)     : 배열 인덱스 접근, 해시맵 조회 → 입력 크기 무관\n"
        "O(log n) : 이진 탐색, BST 탐색 → 매 단계 절반 제거\n"
        "O(n)     : 선형 탐색, 배열 순회\n"
        "O(n log n): 병합 정렬, 힙 정렬\n"
        "O(n²)   : 버블/삽입 정렬, 중첩 루프\n"
        "O(2ⁿ)   : 브루트포스 부분집합 열거\n\n"
        "게임 컨텍스트:\n"
        "매 프레임 O(n²) 충돌 체크 → 오브젝트 100개 시 10,000 비교. Spatial Partition으로 O(n log n)으로 개선.",
        "BigO·시간복잡도·이진탐색·정렬·SpatialPartition",
        "중",
        "최악·평균·최선 복잡도가 다른 알고리즘 예시를 하나 들어라."
    ),
    (
        "콜 스택(Call Stack)이란? 스택 오버플로우는 언제 발생하는가?",
        "콜 스택:\n"
        "함수 호출 시 스택 프레임(리턴 주소, 지역 변수, 매개변수)이 쌓이는 LIFO 구조 메모리.\n\n"
        "스택 오버플로우 발생 원인:\n"
        "① 무한 재귀 (탈출 조건 누락)\n"
        "② 지역 변수로 매우 큰 배열 선언 (int arr[10000000])\n"
        "③ 재귀 깊이가 스택 크기(기본 1~8MB)를 초과\n\n"
        "해결:\n"
        "- 재귀 → 반복문(이터레이티브)으로 변환\n"
        "- 꼬리 재귀 최적화(TCO) — 컴파일러 지원 시\n"
        "- 큰 배열 → 힙 할당 (new/malloc)\n\n"
        "게임: UE5 리커시브 블루프린트 → 스택 오버플로우 실제 발생 사례.",
        "콜스택·스택오버플로우·재귀·스택프레임·TCO",
        "중",
        "재귀 DFS를 스택 자료구조로 변환하는 방법은?"
    ),
    (
        "링크드 리스트 역순(Reverse) 구현을 설명하라.",
        "핵심 아이디어: 포인터 3개(prev, curr, next) 이동.\n\n"
        "코드:\n"
        "Node* prev = nullptr;\n"
        "Node* curr = head;\n"
        "while (curr != nullptr) {\n"
        "    Node* next = curr->next;  // 다음 저장\n"
        "    curr->next = prev;        // 방향 뒤집기\n"
        "    prev = curr;              // prev 전진\n"
        "    curr = next;              // curr 전진\n"
        "}\n"
        "head = prev;\n\n"
        "시간복잡도: O(n), 공간복잡도: O(1)\n\n"
        "면접 팁: 재귀로도 구현 가능하나 스택 오버플로우 위험. 이터레이티브가 안전.",
        "링크드리스트·역순·prev·curr·next·O(n)",
        "중",
        "이중 연결 리스트(Doubly Linked List)를 역순하는 방법은?"
    ),
    (
        "메모리 구조(코드·데이터·힙·스택 영역)를 설명하고 게임에서의 의미는?",
        "Code(Text) 영역: 실행 코드. 읽기 전용.\n"
        "Data 영역: 전역·정적 변수. BSS(초기화 안 된 것) 포함.\n"
        "Heap 영역: 동적 할당(new/malloc). 런타임 크기 결정. 수동 해제 필요.\n"
        "Stack 영역: 지역 변수, 함수 호출 프레임. 자동 해제. 크기 제한 있음.\n\n"
        "게임 관점:\n"
        "- 힙 단편화(fragmentation) → 프레임마다 new/delete 반복 시 발생 → 오브젝트 풀링으로 해결\n"
        "- 스택은 빠르지만 작음(1~8MB) → 큰 버퍼는 반드시 힙에 할당\n"
        "- UE5: UObject는 힙, 지역 FVector는 스택 → GC가 힙만 관리",
        "메모리구조·힙·스택·코드영역·단편화·오브젝트풀",
        "중",
        "힙 단편화를 방지하는 게임 서버에서의 전략을 설명하라."
    ),
    (
        "교착 상태(Deadlock) 발생 4가지 조건과 예방 방법은?",
        "Coffman 4조건 (모두 성립 시 Deadlock):\n"
        "① 상호 배제: 자원을 한 번에 하나만 점유\n"
        "② 점유 대기: 자원 보유 중 다른 자원 대기\n"
        "③ 비선점: 강제로 자원 빼앗을 수 없음\n"
        "④ 순환 대기: A→B→C→A 형태의 대기 사이클\n\n"
        "예방 전략:\n"
        "- 락 순서 고정: 항상 mutex_A → mutex_B 순으로만 획득\n"
        "- std::lock(a, b): 두 락을 원자적으로 동시 획득\n"
        "- 타임아웃: try_lock_for로 일정 시간 후 포기\n"
        "- 락 최소화: 락 보유 시간을 최대한 짧게\n\n"
        "게임 서버: 룸 뮤텍스와 플레이어 뮤텍스를 동시에 잡을 때 순서 고정 필수.",
        "Deadlock·Coffman·순환대기·락순서·std::lock",
        "상",
        "Livelock와 Starvation은 Deadlock과 어떻게 다른가?"
    ),
    (
        "프로세스 간 통신(IPC) 방식의 종류와 게임 서버에서의 활용은?",
        "IPC 종류:\n"
        "- 공유 메모리: 가장 빠름. 동기화 필요 (세마포어/뮤텍스).\n"
        "- 파이프(Pipe): 단방향. 부모-자식 프로세스 간.\n"
        "- 소켓: 네트워크 통신. 프로세스/머신 경계 초월.\n"
        "- 메시지 큐: 비동기 메시지 전달. Redis/Kafka 같은 구조.\n"
        "- 시그널: 비동기 이벤트 알림.\n\n"
        "게임 서버 활용:\n"
        "- LoginServer ↔ LobbyServer ↔ GameServer: TCP 소켓\n"
        "- 세션 토큰 공유: Redis (공유 메모리 역할의 외부 캐시)\n"
        "- 매치메이킹 결과 전달: 메시지 큐",
        "IPC·공유메모리·소켓·메시지큐·Redis·프로세스통신",
        "상",
        "소켓 IPC와 공유 메모리 IPC의 성능 차이를 설명하라."
    ),
],

# ════════════════════════════════════════════════════════════════════
# 03_data_structure.xlsx — 시간복잡도 비교표·트리 보완
# ════════════════════════════════════════════════════════════════════
"03_data_structure": [
    (
        "주요 자료구조의 시간복잡도 비교표를 설명하라.",
        "자료구조       | 접근    | 탐색    | 삽입    | 삭제\n"
        "Array          | O(1)   | O(n)   | O(n)   | O(n)\n"
        "Linked List    | O(n)   | O(n)   | O(1)   | O(1)\n"
        "Stack/Queue    | O(n)   | O(n)   | O(1)   | O(1)\n"
        "Hash Map       | -      | O(1)   | O(1)   | O(1)   ← 평균\n"
        "BST            | -      | O(log n)| O(log n)| O(log n) ← 균형 시\n"
        "Heap           | O(1)*  | O(n)   | O(log n)| O(log n) *최댓값만\n"
        "Graph (인접리스트)| -   | O(V+E) | O(1)   | O(E)\n\n"
        "게임 선택 원칙:\n"
        "- 빠른 조회가 필요 → Hash Map\n"
        "- 우선순위 처리 (데미지 이벤트 큐) → Heap\n"
        "- 순서 보존 + 랜덤 접근 → Array",
        "시간복잡도·Array·HashMap·BST·Heap·Graph",
        "중",
        "Hash Map의 최악 시간복잡도가 O(n)인 이유는?"
    ),
    (
        "트리(Tree)의 종류와 각각의 용도를 설명하라.",
        "이진 트리(Binary Tree): 각 노드 자식 ≤ 2\n"
        "BST: 왼쪽 < 루트 < 오른쪽. 정렬된 탐색. 균형 무너지면 O(n).\n"
        "AVL/Red-Black Tree: 자가 균형. 항상 O(log n). STL map/set 내부.\n"
        "Heap: 최댓값/최솟값 O(1) 접근. 우선순위 큐 구현.\n"
        "Trie: 문자열 접두사 탐색. O(문자열 길이). 자동완성·사전.\n"
        "Segment Tree: 구간 쿼리·업데이트 O(log n). 범위 합/최솟값.\n\n"
        "게임 활용:\n"
        "- BVH(Bounding Volume Hierarchy): 충돌 탐지 O(log n)\n"
        "- Behavior Tree: AI 의사 결정 트리 구조\n"
        "- k-d Tree: 공간 분할, 최근접 탐색",
        "BST·AVL·RedBlack·Heap·Trie·BVH·BehaviorTree",
        "중",
        "Behavior Tree가 일반 트리 자료구조와 다른 점은?"
    ),
    (
        "그래프 탐색(DFS·BFS)의 차이와 게임에서의 활용은?",
        "DFS (깊이 우선 탐색):\n"
        "- 스택(재귀) 사용. 한 방향으로 끝까지 탐색.\n"
        "- 활용: 미로 탈출 경로, 연결 요소 탐색, 위상 정렬\n"
        "- 최단 경로 보장 안 함\n\n"
        "BFS (너비 우선 탐색):\n"
        "- 큐 사용. 현재 레벨을 모두 탐색 후 다음 레벨.\n"
        "- 활용: 최단 경로(가중치 없는 그래프), NavMesh 거리 계산\n"
        "- 가중치 있으면 다익스트라 사용\n\n"
        "게임 활용:\n"
        "- BFS: 타일맵 최단 경로, 시야 범위 계산\n"
        "- DFS: 던전 생성 알고리즘, 퍼즐 상태 탐색\n"
        "- A*: BFS + 휴리스틱 → NavMesh 길찾기",
        "DFS·BFS·최단경로·NavMesh·A*·큐·스택",
        "중",
        "A*가 BFS보다 빠른 이유를 휴리스틱 관점에서 설명하라."
    ),
],

# ════════════════════════════════════════════════════════════════════
# 04_unreal.xlsx — Replication·GC·BT·NavMesh·오브젝트풀링·커리어
# ════════════════════════════════════════════════════════════════════
"04_unreal": [

    # ── Replication ────────────────────────────────────────────────
    (
        "UE5 Replication 구조를 설명하라. UPROPERTY(Replicated)와 ReplicatedUsing의 차이는?",
        "Replication 기본 구조:\n"
        "- 서버가 권위(Authority). 클라이언트는 서버 상태를 수신.\n"
        "- Actor의 bReplicates = true 설정 필요.\n"
        "- GetLifetimeReplicatedProps()에서 복제할 프로퍼티 등록.\n\n"
        "UPROPERTY(Replicated):\n"
        "값이 서버→클라이언트로 전송됨. 클라이언트에서 별도 처리 없음.\n\n"
        "UPROPERTY(ReplicatedUsing=OnRep_함수명):\n"
        "값 복제 + 클라이언트에서 콜백 함수 자동 호출.\n"
        "→ HP 변경 시 UI 업데이트 등 사이드 이펙트 처리에 사용.\n\n"
        "DOREPLIFETIME 조건:\n"
        "DOREPLIFETIME_CONDITION(AMyActor, HP, COND_OwnerOnly) → 소유 클라이언트에만 전송\n\n"
        "NetUpdateFrequency: 초당 복제 횟수. 총알은 높게, 환경은 낮게.",
        "Replication·Replicated·ReplicatedUsing·OnRep·DOREPLIFETIME·Authority",
        "상",
        "서버 RPC(Server RPC)와 Multicast RPC의 차이와 각각의 사용 시점은?"
    ),
    (
        "RPC(Remote Procedure Call) 3종(Server·Client·Multicast)의 차이와 사용 시점은?",
        "Server RPC (UFUNCTION(Server, Reliable/Unreliable)):\n"
        "- 클라이언트가 호출 → 서버에서 실행\n"
        "- 용도: 클라이언트 입력 전달 (이동, 스킬 시전)\n"
        "- WithValidation: 서버에서 입력 유효성 검사 가능\n\n"
        "Client RPC (UFUNCTION(Client, Reliable)):\n"
        "- 서버가 호출 → 특정 클라이언트에서 실행\n"
        "- 용도: 그 클라이언트에게만 피격 피드백, UI 업데이트\n\n"
        "Multicast RPC (UFUNCTION(NetMulticast)):\n"
        "- 서버가 호출 → 모든 클라이언트에서 실행\n"
        "- 용도: 폭발 이펙트, 사운드 등 모두에게 보여야 하는 비주얼\n\n"
        "Reliable vs Unreliable:\n"
        "- Reliable: 재전송 보장. 중요 게임 로직에 사용.\n"
        "- Unreliable: 손실 허용. 이동 업데이트 등 빠른 갱신에 사용.",
        "ServerRPC·ClientRPC·Multicast·Reliable·Unreliable·WithValidation",
        "상",
        "Reliable RPC를 남용하면 어떤 문제가 생기는가?"
    ),

    # ── Garbage Collector ──────────────────────────────────────────
    (
        "UE5 Garbage Collector 동작 방식을 설명하라. UPROPERTY가 GC에 미치는 영향은?",
        "UE5 GC: Mark & Sweep 방식. 주기적으로 실행.\n\n"
        "① Mark 단계:\n"
        "  - GC Root(UGameInstance, UWorld 등)에서 시작\n"
        "  - UPROPERTY로 참조된 UObject를 순회하며 '도달 가능' 표시\n\n"
        "② Sweep 단계:\n"
        "  - 표시되지 않은(도달 불가) UObject를 BeginDestroy() 호출 후 메모리 해제\n\n"
        "UPROPERTY의 역할:\n"
        "- UObject* 멤버를 UPROPERTY로 선언해야 GC가 참조 관계를 추적\n"
        "- UPROPERTY 없이 raw UObject* → GC가 참조 모름 → 수집 대상 → 댕글링\n\n"
        "AddToRoot() / RemoveFromRoot():\n"
        "- GC Root에 직접 추가 → 절대 수집 안 됨. 전역 싱글톤 등에 사용.\n"
        "- 남용 시 메모리 누수.\n\n"
        "TWeakObjectPtr: GC가 수집해도 자동 Null. 순환 참조 방지.",
        "GC·MarkSweep·UPROPERTY·AddToRoot·TWeakObjectPtr·댕글링",
        "상",
        "UPROPERTY 없이 UObject*를 멤버로 들고 있으면 어떤 일이 발생하는가?"
    ),

    # ── Behavior Tree ──────────────────────────────────────────────
    (
        "Behavior Tree 구조(Selector·Sequence·Task·Decorator)를 설명하고 Blackboard 역할은?",
        "BT 노드 종류:\n"
        "Selector(?):\n"
        "  - 자식 노드를 왼쪽부터 실행. 하나라도 Success → 전체 Success.\n"
        "  - 용도: '공격 OR 도망 OR 대기' 중 하나 선택\n\n"
        "Sequence(→):\n"
        "  - 자식 노드를 왼쪽부터 실행. 하나라도 Fail → 전체 Fail.\n"
        "  - 용도: '목표 탐지 AND 이동 AND 공격' 순서 실행\n\n"
        "Task:\n"
        "  - 실제 행동 리프 노드. BTTask_MoveTo, BTTask_PlayAnimation 등.\n\n"
        "Decorator:\n"
        "  - 조건 체크. 노드 실행 전 Guard 역할.\n"
        "  - 예: 체력 < 30% 일 때만 도망 Subtree 활성화\n\n"
        "Blackboard:\n"
        "  - BT 전용 공유 키-값 저장소. AI 상태 데이터 보관.\n"
        "  - TargetActor, PatrolPoint, IsAlerted 등을 Key로 등록.\n"
        "  - Task·Decorator가 Blackboard 값을 읽고 씀.\n\n"
        "BT vs State Machine 선택:\n"
        "  - 상태 수가 적고 전이 단순 → State Machine\n"
        "  - 조건 조합이 복잡하고 재사용 필요 → Behavior Tree",
        "BT·Selector·Sequence·Task·Decorator·Blackboard·StateTree",
        "상",
        "Behavior Tree에서 Abort 기능(Observer Abort)이란 무엇이고 언제 사용하는가?"
    ),

    # ── NavMesh ────────────────────────────────────────────────────
    (
        "UE5 NavMesh 동작 방식과 Dynamic NavMesh 설정 방법은?",
        "NavMesh 기본 동작:\n"
        "- 레벨의 정적 지오메트리를 분석해 이동 가능 영역을 폴리곤 메시로 표현\n"
        "- AI가 MoveToActor/MoveToLocation 호출 시 A* 알고리즘으로 경로 탐색\n"
        "- NavMeshBoundsVolume으로 빌드 범위 지정\n\n"
        "경로 찾기 흐름:\n"
        "AIController.MoveToActor()\n"
        "→ UNavigationSystemV1::FindPathSync()\n"
        "→ A* on NavMesh\n"
        "→ FNavPathSharedPtr (경로 결과)\n\n"
        "Dynamic NavMesh:\n"
        "- ProjectSettings → Navigation → Runtime Generation = Dynamic\n"
        "- 움직이는 장애물 주변 NavMesh 자동 재빌드\n"
        "- 재빌드 비용 있음 → 큰 영역은 Invoker 컴포넌트로 로컬 재빌드\n\n"
        "ProjectPointToNavigation(): 임의 위치를 가장 가까운 NavMesh 위로 투영.",
        "NavMesh·A*·AIController·DynamicNavMesh·Invoker·UNavigationSystem",
        "중",
        "NavMesh가 없는 공중·수중 이동 AI를 구현하려면 어떻게 해야 하는가?"
    ),

    # ── 오브젝트 풀링 ──────────────────────────────────────────────
    (
        "오브젝트 풀링(Object Pooling)이 필요한 이유와 UE5에서의 구현 방법은?",
        "왜 필요한가:\n"
        "SpawnActor: 메모리 할당 + CDO 복사 + 컴포넌트 초기화 → 프레임당 수십 호출 시 히치(hitch) 유발.\n"
        "DestroyActor: GC Mark 트리거 → GC 실행 타이밍에 프레임 드랍.\n"
        "→ 총알·이펙트·적 같이 빈번히 생성·소멸하는 객체에 풀링 필수.\n\n"
        "UE5 구현 패턴:\n"
        "① 초기화 시 N개 SpawnActor → 비활성(SetActorHiddenInGame, SetActorEnableCollision false)\n"
        "② 요청 시 풀에서 꺼내 활성화 (Activate)\n"
        "③ 사용 완료 시 비활성화 후 풀에 반환\n\n"
        "UE5 5.0+: UGameplayStaticsExtended 또는 커스텀 UActorPoolComponent 사용.\n\n"
        "장점: 힙 할당·해제 제거 → 프레임 히치 방지, GC 압력 감소\n"
        "단점: 초기 메모리 선점, 풀 크기 예측 필요",
        "오브젝트풀링·SpawnActor비용·GC압력·히치·SetActorHidden",
        "중",
        "풀 크기를 동적으로 조절하는 전략은? 풀이 고갈되면 어떻게 처리하는가?"
    ),

    # ── AI Perception ──────────────────────────────────────────────
    (
        "UE5 AI Perception (AIPerceptionComponent)의 구조와 Sight·Hearing Sense 설정 방법은?",
        "AIPerceptionComponent:\n"
        "- AIController에 부착. 주변 자극(Stimulus)을 감지.\n"
        "- 감지 시 OnPerceptionUpdated 델리게이트 호출.\n\n"
        "Sight Sense (UAISenseConfig_Sight):\n"
        "- SightRadius: 시야 거리\n"
        "- PeripheralVisionAngleDegrees: 시야각 (양쪽 합산)\n"
        "- LoseSightRadius: 시야 이탈 거리 (Sight보다 크게 설정)\n"
        "- DetectionByAffiliation: 적/아군/중립 필터\n\n"
        "Hearing Sense (UAISenseConfig_Hearing):\n"
        "- HearingRange: 소리 감지 반경\n"
        "- UAIPerceptionSystem::ReportEvent()로 소리 이벤트 발생시킴\n\n"
        "Blackboard 연동:\n"
        "감지된 액터를 Blackboard의 TargetActor 키에 저장\n"
        "→ BT Decorator가 키 존재 여부로 추격/탐색 전환.",
        "AIPerception·SightSense·HearingSense·Blackboard·OnPerceptionUpdated",
        "중",
        "Sight Sense에서 LoseSightRadius를 SightRadius보다 크게 설정하는 이유는?"
    ),

    # ── 월드 크기 관리 ─────────────────────────────────────────────
    (
        "UE5 대형 오픈월드 관리 방법(World Partition·레벨 스트리밍)을 설명하라.",
        "레벨 스트리밍 (전통 방식):\n"
        "- 월드를 여러 서브레벨로 분할\n"
        "- LoadStreamLevel / UnloadStreamLevel으로 런타임 로드/언로드\n"
        "- 수동 분할 → 대형 오픈월드에서 관리 복잡\n\n"
        "World Partition (UE5 신방식):\n"
        "- 월드를 그리드 셀로 자동 분할\n"
        "- 플레이어 주변 셀만 자동 스트리밍\n"
        "- Data Layer: 낮/밤, 이벤트별 콘텐츠 분리 로드\n"
        "- HLOD(Hierarchical LOD): 원거리 셀을 저해상도 메시로 대체\n\n"
        "Nanite + Lumen과 연계:\n"
        "- Nanite: 폴리곤 수 제한 없이 자동 LOD\n"
        "- Lumen: 동적 글로벌 일루미네이션 → 스트리밍 영역에도 적용\n\n"
        "게임 서버 관점: 서버도 같은 셀 기반으로 관심 영역(AOI) 분리.",
        "WorldPartition·레벨스트리밍·HLOD·DataLayer·Nanite·AOI",
        "상",
        "World Partition에서 셀 크기를 결정하는 기준은 무엇인가?"
    ),

    # ── 커리어·개인 ────────────────────────────────────────────────
    (
        "기술 면접에서 '5년 후 커리어 방향'을 묻는 질문에 어떻게 답하는가?",
        "구조: 현재 위치 확인 → 단기 목표 → 장기 방향 → 회사와의 연결\n\n"
        "예시 답변 (게임 서버 지망):\n"
        "'현재 UE5 클라이언트와 C++ 서버 양쪽을 경험했습니다.\n"
        "3년 안에는 실시간 게임 서버의 네트워크 동기화와 서버 아키텍처 설계를 깊이 있게 익히고,\n"
        "5년에는 서버 최적화와 대규모 동접자 처리 경험을 쌓아\n"
        "클라이언트와 서버 양쪽을 이해하는 게임 서버 리드 엔지니어로 성장하고 싶습니다.\n"
        "MU Online처럼 오랜 라이브 서비스를 운영하는 환경에서 실전 경험을 쌓는 것이 이 목표에 가장 빠른 길이라고 생각합니다.'\n\n"
        "핵심: 추상적 포부보다 구체적 기술 성장 경로 + 지원 회사와의 연결.",
        "커리어방향·5년후·기술성장·라이브서비스·리드엔지니어",
        "중",
        "게임 업계에서 클라이언트와 서버를 모두 할 수 있는 것의 장단점은?"
    ),
    (
        "기술 면접에서 '장점·단점'을 묻는 질문에 어떻게 답하는가?",
        "장점 — 기술 직군에 맞는 구체적 근거 포함:\n"
        "'문제의 근본 원인을 찾을 때까지 파고드는 성향이 있습니다.\n"
        "InfinityServer에서 틱 오버 현상이 간헐적으로 발생했을 때,\n"
        "단순히 sleep 값을 조정하는 것이 아니라 steady_clock 기반 절대 시점 슬립으로\n"
        "누적 오차 구조 자체를 바꿔 근본적으로 해결한 것이 이 성향의 예입니다.'\n\n"
        "단점 — 실제 단점 + 개선 노력:\n"
        "'새로운 구조를 설계할 때 지나치게 완전한 설계를 추구하다 초기 구현이 느려지는 경향이 있습니다.\n"
        "이를 보완하기 위해 먼저 동작하는 최소 구현을 완성한 뒤 리팩토링하는 방식을 의식적으로 연습 중입니다.'\n\n"
        "금지: 단점을 장점으로 포장하는 답변 ('완벽주의가 단점'). 면접관이 꿰뚫음.",
        "장점·단점·자기인식·구체적근거·개선노력",
        "중",
        "이전 팀에서 본인의 어떤 점이 동료에게 불편하게 느껴졌을 수 있는가?"
    ),
    (
        "게임·메타버스 도메인에 대한 본인의 관점을 묻는 질문에 어떻게 답하는가?",
        "구조: 도메인 현황 인식 → 기술 관점 → 본인 기여 방향\n\n"
        "예시 (게임 서버 포지션):\n"
        "'게임은 실시간 동기화, 물리 시뮬레이션, 대규모 동접자 처리가 동시에 요구되는\n"
        "기술적으로 가장 까다로운 서버 환경 중 하나라고 생각합니다.\n"
        "특히 MU Online처럼 PvP 중심의 라이브 게임은\n"
        "수십 년간 서비스를 유지하면서 신규 콘텐츠를 계속 추가해야 하는데,\n"
        "이를 위한 서버 안정성과 확장성 설계가 핵심 과제라고 봅니다.\n"
        "저는 Authoritative 서버 설계와 Dead Reckoning 같은 기술로 이 과제에 기여할 수 있습니다.'\n\n"
        "주의: '게임이 좋아서'만 말하면 0점. 기술적 흥미와 기여 방향이 핵심.",
        "도메인관점·라이브서비스·기술기여·PvP·안정성",
        "중",
        "10년 후 게임 서버 기술이 어떻게 변할 것이라고 생각하는가?"
    ),

    # ── 서버·네트워크 ──────────────────────────────────────────────
    (
        "Dedicated Server 경험을 설명하라. Listen Server와의 차이는?",
        "Listen Server:\n"
        "- 플레이어 한 명이 서버 겸 클라이언트 역할\n"
        "- 호스트 유저에게 유리한 레이턴시 불균형 발생\n"
        "- 소규모 Co-op에 적합\n\n"
        "Dedicated Server:\n"
        "- 게임 로직만 실행하는 독립 프로세스. 렌더링 없음.\n"
        "- 모든 클라이언트가 동등한 레이턴시 → 공정한 환경\n"
        "- 치팅 방지: 서버가 권위(Authority)\n\n"
        "InfinityServer 구현 경험:\n"
        "- C++ Winsock2 TCP, LoginServer·LobbyServer·GameServer 3계층\n"
        "- 고정 틱레이트(20Hz), Authoritative 물리, Dead Reckoning\n"
        "- 플레이어당 1스레드 모델 → 개선점: IOCP 비동기 I/O\n\n"
        "UE5 DS 빌드: -server 플래그로 쿠킹, WITH_SERVER_CODE 매크로로 서버 전용 코드 분리.",
        "DedicatedServer·ListenServer·Authority·IOCP·WithServerCode",
        "상",
        "UE5 Dedicated Server에서 클라이언트 입력 검증(Validation)은 어떻게 구현하는가?"
    ),
    (
        "네트워크 동기화 이슈(고지연·패킷 손실)를 어떻게 해결했는가?",
        "Dead Reckoning (위치 예측):\n"
        "- 마지막 권위 위치 + 속도 × 경과시간으로 현재 위치 예측\n"
        "- 50ms 공백 구간 동안 클라이언트가 부드럽게 위치 보간\n"
        "- 오차가 임계값(50 units) 초과 시에만 보정 패킷 전송 → 대역폭 절감\n\n"
        "Lag Compensation (지연 보상):\n"
        "- 서버가 최근 64틱 스냅샷을 링 버퍼에 보존\n"
        "- 클라이언트가 스킬 사용 시 틱 번호 함께 전송\n"
        "- 서버가 해당 틱의 피격 대상 위치를 재현 → 공정한 히트 판정\n\n"
        "Delta 동기화:\n"
        "- dirty 비트마스크로 변경된 필드만 전송\n"
        "- 정지한 플레이어는 전송량 0\n\n"
        "패킷 손실 복구:\n"
        "- 5초 주기 Full Snapshot 전송으로 누락 상태 자동 재정렬",
        "DeadReckoning·LagCompensation·DeltaSync·DirtyFlag·Snapshot",
        "상",
        "Dead Reckoning의 예측 오차가 커지는 상황은 언제이고 어떻게 처리하는가?"
    ),
    (
        "NetRole과 Authority 구분을 설명하라. HasAuthority()는 언제 사용하는가?",
        "NetRole 종류:\n"
        "ROLE_Authority:\n"
        "  - 서버에서 실행되는 Actor. 게임 상태의 진실 소유자.\n"
        "ROLE_AutonomousProxy:\n"
        "  - 플레이어 본인이 제어하는 클라이언트 측 Actor. 입력 선처리 가능.\n"
        "ROLE_SimulatedProxy:\n"
        "  - 다른 플레이어의 클라이언트 측 복제본. 서버 값을 수신해 표시만 함.\n\n"
        "HasAuthority() 사용 시점:\n"
        "- 게임 로직(데미지, 아이템 지급, 스폰)은 반드시 서버에서만 실행\n"
        "- if (HasAuthority()) { ApplyDamage(); } → 클라이언트에서 실행 방지\n\n"
        "IsLocallyControlled() 사용 시점:\n"
        "- 입력 처리, 카메라, 사운드 등 로컬 플레이어에만 필요한 로직.",
        "NetRole·Authority·AutonomousProxy·SimulatedProxy·HasAuthority",
        "상",
        "클라이언트에서 HasAuthority()가 true인 경우는 어떤 상황인가?"
    ),
],
}

# ─── 파일별 설정 ─────────────────────────────────────────────────
FILE_CFG = {
    "01_cpp": {
        "file": "01_cpp.xlsx",
        "header_bg": "1A3A5C",
        "row_q_bg":  "D6E4F7",
        "row_a_bg":  "EAF4FF",
    },
    "02_cs": {
        "file": "02_cs.xlsx",
        "header_bg": "1A4731",
        "row_q_bg":  "D4EDDA",
        "row_a_bg":  "EBF7EE",
    },
    "03_data_structure": {
        "file": "03_data_structure.xlsx",
        "header_bg": "5C2A00",
        "row_q_bg":  "FDEBD0",
        "row_a_bg":  "FEF5E7",
    },
    "04_unreal": {
        "file": "04_unreal.xlsx",
        "header_bg": "3B1F5E",
        "row_q_bg":  "E8D5F5",
        "row_a_bg":  "F3E8FF",
    },
}

# ─── 병합 실행 ────────────────────────────────────────────────────
for cat_key, new_items in NEW_CARDS.items():
    cfg  = FILE_CFG[cat_key]
    path = os.path.join(MASTER, cfg["file"])

    wb = openpyxl.load_workbook(path)

    # ── 플래시카드 시트 ──
    ws = wb["플래시카드"]
    hbg = cfg["header_bg"]
    qbg = cfg["row_q_bg"]
    abg = cfg["row_a_bg"]

    # 기존 질문 집합 (중복 방지)
    existing_qs = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]: existing_qs.add(str(row[1]).strip())

    start_row = ws.max_row + 1
    added = 0

    for item in new_items:
        q, a, kw, lvl, tail = item
        if q.strip() in existing_qs:
            print(f"  [SKIP 중복] {q[:40]}")
            continue

        ri = start_row + added
        bg_q = qbg if ri % 2 == 0 else abg
        bg_a = abg if ri % 2 == 0 else qbg

        vals = [str(ri - 1), q, a, kw, lvl, tail]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            is_q = (ci == 2)
            c.fill      = fill(bg_q if is_q else bg_a)
            c.font      = font(bold=is_q, size=10 if not is_q else 11)
            c.alignment = align("left", "center")
            c.border    = thin_border()

        a_len = len(a)
        h = 40 if a_len < 100 else (65 if a_len < 250 else 100)
        ws.row_dimensions[ri].height = h

        existing_qs.add(q.strip())
        added += 1

    # ── 빠른복습 시트 동기화 ──
    ws_q = wb["빠른복습"]
    alt  = ["F2F3F4", "EAECEE"]
    qr_start = ws_q.max_row + 1

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 빠른복습은 전체 재기록 (번호 순서 맞추기)
    # 기존 행 수 확인
    existing_quick = ws_q.max_row - 1  # 헤더 제외

    # 새로 추가된 카드만 빠른복습에 추가
    for i in range(added):
        src_row_idx = start_row + i  # 플래시카드 시트 행 번호
        src_row = ws.iter_rows(min_row=src_row_idx, max_row=src_row_idx, values_only=True)
        src = list(src_row)[0]
        q_text = str(src[1]) if src[1] else ""
        kw_text = str(src[3]) if src[3] else ""

        ri_q = qr_start + i
        bg = alt[ri_q % 2]
        num_val = str(ri_q - 1)
        for ci, val in enumerate([num_val, q_text, kw_text], 1):
            c = ws_q.cell(row=ri_q, column=ci, value=val)
            c.font      = font(bold=(ci == 2), size=10)
            c.fill      = fill(bg)
            c.alignment = align("center" if ci in (1, 3) else "left", "center")
            c.border    = thin_border()
        ws_q.row_dimensions[ri_q].height = 28

    wb.save(path)
    total = ws.max_row - 1
    print(f"[OK] {cfg['file']}  +{added}개 추가 → 총 {total}개")

print("\n추가 완료.")
